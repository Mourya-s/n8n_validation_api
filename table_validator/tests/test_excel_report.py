"""
Tests for reports/excel_report.py's row-building logic. Focused on the
missing_tables/extra_tables labeling bug: a table present in target but
absent from source must be labeled EXTRA_IN_TARGET, not lumped in with
MISSING_FROM_TARGET (which means the opposite - present in source,
absent from target).
"""

from __future__ import annotations

import pytest

from table_validator.models import (
    CatalogValidationResponse,
    ColumnValidationResult,
    DataValidationResult,
    RowHashMismatch,
    RowMismatchDetail,
    SchemaValidationResult,
    TableValidationResult,
    ValidationStatus,
    ValidationTier,
)
from table_validator.reports.excel_report import (
    CATEGORY_SUMMARY_HEADERS,
    MISMATCH_HEADERS,
    ROW_HASH_HEADERS,
    SUGGESTION_HEADERS,
    TABLE_HEADERS,
    ColumnLenStats,
    MismatchCategorySummary,
    _build_mismatch_category_summary,
    _build_mismatch_rows,
    _build_row_hash_rows,
    _build_suggestion_rows,
    _build_table_rows,
    _classify_all_mismatches,
    build_category_sentences,
    build_column_len_stats,
    generate_excel_report,
)

_SOURCE_TABLE_COL = 1
_STATUS_COL = 4


def _response(schema_result: SchemaValidationResult) -> CatalogValidationResponse:
    return CatalogValidationResponse(
        source_catalog="src",
        target_catalog="tgt",
        status=ValidationStatus.FAIL,
        schemas=[schema_result],
    )


def test_missing_tables_rendered_as_missing_from_target():
    schema = SchemaValidationResult(
        schema_name="bronze",
        status=ValidationStatus.FAIL,
        missing_tables=["only_in_source"],
        extra_tables=[],
        tables=[],
    )
    rows = _build_table_rows(_response(schema))

    assert len(rows) == 1
    assert rows[0][_SOURCE_TABLE_COL] == "only_in_source"
    assert rows[0][_STATUS_COL] == "MISSING_FROM_TARGET"


def test_extra_tables_rendered_as_extra_in_target_not_missing():
    """Regression test: a table present in target but with no matching
    source (e.g. a Databricks table with no matching blob) must be
    labeled EXTRA_IN_TARGET - previously extra_tables was never rendered
    at all, and blob_discovery.py separately mislabeled these as
    MISSING_FROM_TARGET by putting them in the wrong field."""
    schema = SchemaValidationResult(
        schema_name="bronze",
        status=ValidationStatus.FAIL,
        missing_tables=[],
        extra_tables=["only_in_target"],
        tables=[],
    )
    rows = _build_table_rows(_response(schema))

    assert len(rows) == 1
    assert rows[0][_SOURCE_TABLE_COL] == "only_in_target"
    assert rows[0][_STATUS_COL] == "EXTRA_IN_TARGET"


def test_both_missing_and_extra_tables_rendered_distinctly():
    schema = SchemaValidationResult(
        schema_name="bronze",
        status=ValidationStatus.FAIL,
        missing_tables=["only_in_source"],
        extra_tables=["only_in_target"],
        tables=[],
    )
    rows = _build_table_rows(_response(schema))

    statuses_by_table = {r[_SOURCE_TABLE_COL]: r[_STATUS_COL] for r in rows}
    assert statuses_by_table == {
        "only_in_source": "MISSING_FROM_TARGET",
        "only_in_target": "EXTRA_IN_TARGET",
    }


def test_matched_table_and_extra_table_both_appear():
    matched = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.PASS,
    )
    schema = SchemaValidationResult(
        schema_name="bronze",
        status=ValidationStatus.FAIL,
        missing_tables=[],
        extra_tables=["orphan_table"],
        tables=[matched],
    )
    rows = _build_table_rows(_response(schema))

    statuses_by_table = {r[_SOURCE_TABLE_COL]: r[_STATUS_COL] for r in rows}
    assert statuses_by_table["customers"] == "PASS"
    assert statuses_by_table["orphan_table"] == "EXTRA_IN_TARGET"


def test_all_rows_match_header_width_including_missing_and_extra():
    """Missing/extra table rows are built as a fixed-length filler list
    ("" * N) rather than field-by-field - a new TABLE_HEADERS column
    (e.g. Tier Reached) must have that filler count updated in lockstep,
    or every missing/extra row silently misaligns against the header."""
    matched = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.PASS,
        tier_reached=ValidationTier.FINGERPRINT,
    )
    schema = SchemaValidationResult(
        schema_name="bronze",
        status=ValidationStatus.FAIL,
        missing_tables=["gone"],
        extra_tables=["orphan"],
        tables=[matched],
    )
    rows = _build_table_rows(_response(schema))

    for row in rows:
        assert len(row) == len(TABLE_HEADERS)


def test_tier_reached_populated_for_matched_table():
    matched = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.PASS,
        tier_reached=ValidationTier.FINGERPRINT,
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.PASS, tables=[matched],
    )
    rows = _build_table_rows(_response(schema))

    tier_col = TABLE_HEADERS.index("Tier Reached")
    assert rows[0][tier_col] == "FINGERPRINT"


def test_schema_blocking_gets_top_priority_suggestion():
    blocked = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        schema_blocking=True, tier_reached=ValidationTier.SCHEMA_BLOCKED,
        missing_columns=["age"], columns_status=ValidationStatus.FAIL,
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[blocked],
    )
    rows = _build_suggestion_rows(_response(schema))

    issue_types = [r[3] for r in rows]
    assert "Blocked at Schema Check" in issue_types
    assert issue_types[0] == "Blocked at Schema Check"


def test_partition_column_shows_bucket_summary_when_partitioned():
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        tier_reached=ValidationTier.ROW_HASH,
        partitioned=True, partition_column="region",
        partition_buckets_total=10, partition_buckets_culprit=2,
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table],
    )
    rows = _build_table_rows(_response(schema))

    partition_col = TABLE_HEADERS.index("Partition")
    assert rows[0][partition_col] == "region (2/10 buckets differed)"


def test_partition_column_shows_skip_reason_when_not_partitioned():
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        tier_reached=ValidationTier.ROW_HASH,
        partitioned=False, partition_skip_reason="user declined or non-interactive run",
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table],
    )
    rows = _build_table_rows(_response(schema))

    partition_col = TABLE_HEADERS.index("Partition")
    assert rows[0][partition_col] == "not partitioned (user declined or non-interactive run)"


def test_partition_column_blank_when_not_offered():
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.PASS,
        tier_reached=ValidationTier.FINGERPRINT,
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.PASS, tables=[table],
    )
    rows = _build_table_rows(_response(schema))

    partition_col = TABLE_HEADERS.index("Partition")
    assert rows[0][partition_col] == ""


def test_row_hash_mismatches_sheet_includes_partition_bucket_column():
    mismatch = RowHashMismatch(
        primary_key="42", source_hash="aaa", target_hash="bbb",
        status="MISMATCH", partition_bucket="west",
    )
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        data=DataValidationResult(
            mode="STATISTICS", status=ValidationStatus.FAIL,
            row_hash_mismatches=[mismatch],
        ),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table],
    )
    rows = _build_row_hash_rows(_response(schema))

    bucket_col = ROW_HASH_HEADERS.index("Partition Bucket")
    assert rows[0][bucket_col] == "west"


def test_row_hash_mismatches_sheet_blank_bucket_when_unpartitioned():
    mismatch = RowHashMismatch(
        primary_key="42", source_hash="aaa", target_hash="bbb", status="MISMATCH",
    )
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        data=DataValidationResult(
            mode="STATISTICS", status=ValidationStatus.FAIL,
            row_hash_mismatches=[mismatch],
        ),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table],
    )
    rows = _build_row_hash_rows(_response(schema))

    bucket_col = ROW_HASH_HEADERS.index("Partition Bucket")
    assert rows[0][bucket_col] == ""


def test_mismatch_headers_include_verified_column():
    assert MISMATCH_HEADERS[-1] == "Verified"
    assert len(MISMATCH_HEADERS) == 9


def test_mismatch_rows_show_yes_for_verified_detail():
    detail = RowMismatchDetail(
        schema_name="bronze", table="customers",
        primary_key={"id": 2}, mismatch_column="name",
        source_value="old", target_value="new",
        verified=True,
    )
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        data=DataValidationResult(
            mode="STATISTICS", status=ValidationStatus.FAIL,
            sample_changed_detail=[detail],
        ),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table],
    )
    rows = _build_mismatch_rows(_response(schema))

    verified_col = MISMATCH_HEADERS.index("Verified")
    assert rows[0][verified_col] == "Yes"


def test_mismatch_rows_label_row_number_fallback_detail_as_unverified():
    detail = RowMismatchDetail(
        schema_name="bronze", table="customers",
        primary_key={"row_number": 2}, mismatch_column="name",
        source_value="old", target_value="new",
        verified=False,
    )
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        data=DataValidationResult(
            mode="STATISTICS", status=ValidationStatus.FAIL,
            sample_changed_detail=[detail],
        ),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table],
    )
    rows = _build_mismatch_rows(_response(schema))

    verified_col = MISMATCH_HEADERS.index("Verified")
    assert rows[0][verified_col] == "No (row-number, unverified)"


def test_suggestion_wording_uses_row_position_for_row_number_fallback():
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        data=DataValidationResult(
            mode="STATISTICS", status=ValidationStatus.FAIL,
            key_columns=["row_number"],
            row_hash_mismatch_count=2,
            row_hash_mismatch_percentage=0.5,
        ),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table],
    )
    rows = _build_suggestion_rows(_response(schema))

    row_mismatch_rows = [r for r in rows if r[3] == "Row Data Mismatch"]
    assert len(row_mismatch_rows) == 1
    suggestion_text = row_mismatch_rows[0][4]
    assert "row position" in suggestion_text
    assert "best-effort" in suggestion_text
    assert "unverified" in suggestion_text


def test_suggestion_wording_uses_primary_key_for_real_key_path():
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        data=DataValidationResult(
            mode="STATISTICS", status=ValidationStatus.FAIL,
            key_columns=["id"],
            row_hash_mismatch_count=2,
            row_hash_mismatch_percentage=0.5,
        ),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table],
    )
    rows = _build_suggestion_rows(_response(schema))

    row_mismatch_rows = [r for r in rows if r[3] == "Row Data Mismatch"]
    assert len(row_mismatch_rows) == 1
    suggestion_text = row_mismatch_rows[0][4]
    assert "by primary key" in suggestion_text
    assert "row position" not in suggestion_text


# ---------------------------------------------------------------------------
# Mismatch Count/% (Table Validation sheet): the tiered fail-fast funnel
# never populates source_only_rows/target_only_rows/changed_rows (those
# only ever came from the legacy FULL-mode EXCEPT/hash-join path) - these
# two columns must fall back to the row-hash comparison's own count/%
# instead of staying permanently blank for every tiered-pipeline table.
# ---------------------------------------------------------------------------
def test_mismatch_count_falls_back_to_row_hash_count_for_tiered_pipeline():
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        row_count_source=100, row_count_target=100,
        data=DataValidationResult(
            mode="STATISTICS", status=ValidationStatus.FAIL,
            row_hash_mismatch_count=3,
            row_hash_mismatch_percentage=3.0,
        ),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table],
    )
    rows = _build_table_rows(_response(schema))

    mismatch_count_col = TABLE_HEADERS.index("Mismatch Count")
    mismatch_pct_col = TABLE_HEADERS.index("Mismatch %")
    assert rows[0][mismatch_count_col] == 3
    assert rows[0][mismatch_pct_col] == "3.00%"


def test_mismatch_count_prefers_legacy_fields_when_present():
    """A legacy FULL-mode result (source_only_rows/target_only_rows/
    changed_rows populated) must still take priority over the row-hash
    fallback - the fallback only applies when those are all None."""
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        row_count_source=100, row_count_target=100,
        data=DataValidationResult(
            mode="FULL", status=ValidationStatus.FAIL,
            source_only_rows=1, target_only_rows=0, changed_rows=2,
            row_hash_mismatch_count=99,  # must be ignored here
            row_hash_mismatch_percentage=99.0,
        ),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table],
    )
    rows = _build_table_rows(_response(schema))

    mismatch_count_col = TABLE_HEADERS.index("Mismatch Count")
    mismatch_pct_col = TABLE_HEADERS.index("Mismatch %")
    assert rows[0][mismatch_count_col] == 3
    assert rows[0][mismatch_pct_col] == "3.00%"


def test_mismatch_count_blank_when_no_mismatches_found_at_all():
    """A clean PASS table (row_hash_mismatch_count=0, the pydantic
    default) must not show a misleading fallback count - stay blank."""
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.PASS,
        row_count_source=100, row_count_target=100,
        data=DataValidationResult(mode="STATISTICS", status=ValidationStatus.PASS),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.PASS, tables=[table],
    )
    rows = _build_table_rows(_response(schema))

    mismatch_count_col = TABLE_HEADERS.index("Mismatch Count")
    mismatch_pct_col = TABLE_HEADERS.index("Mismatch %")
    assert rows[0][mismatch_count_col] == ""
    assert rows[0][mismatch_pct_col] == ""


# ---------------------------------------------------------------------------
# generate_excel_report(): enabled_validations gates whole sheets - "row"
# gates Data Mismatches/Row Hash Mismatches, "column" gates Column
# Validation. Deselecting either must remove the sheet entirely, not just
# leave it empty, so a user who unchecks "row" doesn't see a stray sheet
# that looks like "no mismatches found" when row-level checks never ran.
# ---------------------------------------------------------------------------
def _table_with_a_column(**overrides) -> TableValidationResult:
    defaults = dict(
        schema_name="bronze", table="customers", status=ValidationStatus.PASS,
        exists_in_source=True, exists_in_target=True,
        columns=[ColumnValidationResult(column="id", status=ValidationStatus.PASS)],
    )
    defaults.update(overrides)
    return TableValidationResult(**defaults)


def test_all_validations_enabled_includes_every_sheet(tmp_path):
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.PASS,
        tables=[_table_with_a_column()],
    )
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(_response(schema), str(output_path), enabled_validations=None)

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    assert "Column Validation" in wb.sheetnames
    assert "Data Mismatches" in wb.sheetnames
    assert "Row Hash Mismatches" in wb.sheetnames


def test_deselecting_row_omits_data_and_row_hash_mismatch_sheets(tmp_path):
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.PASS,
        tables=[_table_with_a_column()],
    )
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(
        _response(schema), str(output_path),
        enabled_validations={"catalog", "schema", "column"},
    )

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    assert "Column Validation" in wb.sheetnames
    assert "Data Mismatches" not in wb.sheetnames
    assert "Row Hash Mismatches" not in wb.sheetnames


def test_deselecting_column_omits_column_validation_sheet(tmp_path):
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.PASS,
        tables=[_table_with_a_column()],
    )
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(
        _response(schema), str(output_path),
        enabled_validations={"catalog", "schema", "row"},
    )

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    assert "Column Validation" not in wb.sheetnames
    # row is still enabled, so these sheets stay (even if empty of rows).
    assert "Data Mismatches" in wb.sheetnames
    assert "Row Hash Mismatches" in wb.sheetnames


def test_deselecting_both_column_and_row_omits_all_three_sheets(tmp_path):
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.PASS,
        tables=[_table_with_a_column()],
    )
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(
        _response(schema), str(output_path),
        enabled_validations={"catalog", "schema"},
    )

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    assert "Column Validation" not in wb.sheetnames
    assert "Data Mismatches" not in wb.sheetnames
    assert "Row Hash Mismatches" not in wb.sheetnames
    # Summary/Table Validation/Suggestions are never gated by any single
    # validation type - always present regardless of selection.
    assert "Summary" in wb.sheetnames
    assert "Table Validation" in wb.sheetnames
    assert "Suggestions" in wb.sheetnames


# ---------------------------------------------------------------------------
# _classify_all_mismatches() / _build_mismatch_category_summary(): phase 2
# of the mismatch-categorization feature - classify every RowMismatchDetail
# already collected by the validator, then roll the results up into one
# MismatchCategorySummary per category. Pure in-memory computation; no
# Excel writing, no new queries.
# ---------------------------------------------------------------------------
def _detail(table="customers", column="name", source_value="old", target_value="new", **overrides):
    defaults = dict(
        schema_name="bronze", table=table,
        primary_key={"id": 1}, mismatch_column=column,
        source_value=source_value, target_value=target_value,
    )
    defaults.update(overrides)
    return RowMismatchDetail(**defaults)


def _response_with_details(*details: RowMismatchDetail) -> CatalogValidationResponse:
    """Bundles every given detail onto a single table/schema, for tests
    that only care about the flat classification/aggregation logic, not
    the schema/table structure itself."""
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        data=DataValidationResult(
            mode="FULL", status=ValidationStatus.FAIL,
            sample_changed_detail=list(details),
        ),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table],
    )
    return _response(schema)


def test_classify_all_mismatches_sets_category_on_each_detail_in_place():
    detail = _detail(source_value="ACTIVE", target_value="active")
    response = _response_with_details(detail)

    classified = _classify_all_mismatches(response)

    assert len(classified) == 1
    assert classified[0] is detail
    assert detail.mismatch_category == "CASE_DIFFERENCE"


def test_classify_all_mismatches_returns_flat_list_across_multiple_tables():
    detail_a = _detail(table="customers", source_value=None, target_value="x")
    detail_b = _detail(table="orders", source_value="1,000", target_value="1000")
    table_a = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        data=DataValidationResult(
            mode="FULL", status=ValidationStatus.FAIL, sample_changed_detail=[detail_a],
        ),
    )
    table_b = TableValidationResult(
        schema_name="bronze", table="orders", status=ValidationStatus.FAIL,
        data=DataValidationResult(
            mode="FULL", status=ValidationStatus.FAIL, sample_changed_detail=[detail_b],
        ),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table_a, table_b],
    )
    response = _response(schema)

    classified = _classify_all_mismatches(response)

    assert {d.mismatch_category for d in classified} == {"NULL_MISMATCH", "FORMATTING_DIFF"}


def test_classify_all_mismatches_skips_tables_with_no_data_result():
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.PASS,
        data=None,
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.PASS, tables=[table],
    )
    assert _classify_all_mismatches(_response(schema)) == []


def test_mismatch_category_summary_empty_when_no_mismatches():
    schema = SchemaValidationResult(schema_name="bronze", status=ValidationStatus.PASS, tables=[])
    assert _build_mismatch_category_summary(_response(schema)) == []


def test_mismatch_category_summary_counts_and_percentages():
    response = _response_with_details(
        _detail(column="name", source_value="ACTIVE", target_value="active"),
        _detail(column="name", source_value="PENDING", target_value="pending"),
        _detail(column="status", source_value=None, target_value="x"),
    )

    summary = _build_mismatch_category_summary(response)

    assert isinstance(summary[0], MismatchCategorySummary)
    by_category = {s.category: s for s in summary}
    assert by_category["CASE_DIFFERENCE"].count == 2
    assert by_category["CASE_DIFFERENCE"].pct == pytest.approx(66.67, abs=0.01)
    assert by_category["NULL_MISMATCH"].count == 1
    assert by_category["NULL_MISMATCH"].pct == pytest.approx(33.33, abs=0.01)
    # Percentages are over the TOTAL mismatch count across all categories.
    assert sum(s.count for s in summary) == 3


def test_mismatch_category_summary_identifies_top_table_and_column():
    detail_a = _detail(table="customers", column="name", source_value="ACTIVE", target_value="active")
    detail_b = _detail(table="customers", column="status", source_value="PENDING", target_value="pending")
    detail_c = _detail(table="orders", column="name", source_value="DONE", target_value="done")
    table_customers = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        data=DataValidationResult(
            mode="FULL", status=ValidationStatus.FAIL,
            sample_changed_detail=[detail_a, detail_b],
        ),
    )
    table_orders = TableValidationResult(
        schema_name="bronze", table="orders", status=ValidationStatus.FAIL,
        data=DataValidationResult(
            mode="FULL", status=ValidationStatus.FAIL, sample_changed_detail=[detail_c],
        ),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table_customers, table_orders],
    )
    response = _response(schema)

    summary = _build_mismatch_category_summary(response)

    assert len(summary) == 1
    case_diff = summary[0]
    assert case_diff.category == "CASE_DIFFERENCE"
    assert case_diff.count == 3
    # "customers" has 2 of the 3 CASE_DIFFERENCE mismatches -> top table.
    assert case_diff.top_table == "customers"
    # "name" has 2 of the 3 CASE_DIFFERENCE mismatches -> top column.
    assert case_diff.top_column == "name"


def test_mismatch_category_summary_sorted_most_common_first():
    response = _response_with_details(
        _detail(source_value=None, target_value="x"),        # NULL_MISMATCH
        _detail(source_value="ACTIVE", target_value="active"),  # CASE_DIFFERENCE
        _detail(source_value="PENDING", target_value="pending"),  # CASE_DIFFERENCE
        _detail(source_value="DONE", target_value="done"),  # CASE_DIFFERENCE
    )

    summary = _build_mismatch_category_summary(response)

    assert [s.category for s in summary] == ["CASE_DIFFERENCE", "NULL_MISMATCH"]
    assert summary[0].count == 3
    assert summary[1].count == 1


def test_mismatch_category_summary_uses_canonical_column_name_when_renamed():
    """source_mismatch_column is only set when column_map renamed the
    column - mismatch_column (the canonical/target name) is what every
    other sheet groups by, so the summary must use it too, not the
    source-side spelling."""
    detail = _detail(
        column="customer_id", source_value="ACTIVE", target_value="active",
    )
    detail.source_mismatch_column = "cust_id"
    response = _response_with_details(detail)

    summary = _build_mismatch_category_summary(response)

    assert summary[0].top_column == "customer_id"


# ---------------------------------------------------------------------------
# build_column_len_stats() / build_category_sentences(): phase 3 - one
# honest, specific English sentence per category present in a summary
# list. No fabricated numbers (no confidence scores, no invented decimal-
# place counts) - every value in a sentence traces back to something
# Phase 1/2 actually computed.
# ---------------------------------------------------------------------------
def test_column_len_stats_averages_only_string_truncation_mismatches():
    details = [
        _detail(column="cust_name", source_value="John Doe", target_value="John Do"),   # 8 -> 7
        _detail(column="cust_name", source_value="Jane Smith", target_value="Jane Smit"),  # 10 -> 9
        # Not truncation (case difference) - must NOT be counted here.
        _detail(column="cust_name", source_value="ACTIVE", target_value="active"),
    ]
    response = _response_with_details(*details)
    classified = _classify_all_mismatches(response)

    stats = build_column_len_stats(classified)

    assert stats["cust_name"].avg_source_len == pytest.approx(9.0)
    assert stats["cust_name"].avg_target_len == pytest.approx(8.0)


def test_column_len_stats_empty_when_no_truncation_mismatches():
    details = [_detail(source_value="ACTIVE", target_value="active")]
    response = _response_with_details(*details)
    classified = _classify_all_mismatches(response)

    assert build_column_len_stats(classified) == {}


def test_column_len_stats_skips_non_string_values():
    """STRING_TRUNCATION only ever fires for two actual strings (per
    mismatch_classifier's own rule 2), but this guards independently
    against a caller handing in an unclassified/mislabeled detail whose
    values aren't strings - len() must never be silently applied to a
    non-string via str() coercion."""
    detail = _detail(source_value="abc", target_value="ab")
    detail.mismatch_category = "STRING_TRUNCATION"  # forced, not real classification
    weird_detail = _detail(source_value=123, target_value=12)
    weird_detail.mismatch_category = "STRING_TRUNCATION"  # forced - values aren't strings

    stats = build_column_len_stats([detail, weird_detail])

    assert stats["name"].avg_source_len == pytest.approx(3.0)
    assert stats["name"].avg_target_len == pytest.approx(2.0)


def test_category_sentence_null_mismatch():
    summaries = [
        MismatchCategorySummary(
            category="NULL_MISMATCH", count=5, pct=50.0,
            top_table="customers", top_column="status",
        ),
    ]
    sentences = build_category_sentences(summaries, {})

    assert sentences == [
        "5 nulls mismatched in status — check nullable constraint "
        "differences between source and target."
    ]


def test_category_sentence_string_truncation_with_length_stats():
    summaries = [
        MismatchCategorySummary(
            category="STRING_TRUNCATION", count=2, pct=50.0,
            top_table="customers", top_column="cust_name",
        ),
        MismatchCategorySummary(
            category="CASE_DIFFERENCE", count=2, pct=50.0,
            top_table="customers", top_column="status",
        ),
    ]
    column_stats = {"cust_name": ColumnLenStats(avg_source_len=9.0, avg_target_len=8.0)}

    sentences = build_category_sentences(summaries, column_stats)

    # Total (4) comes from summing every summary's count, not just this
    # category's - "2 of 4", not "2 of 2".
    assert sentences[0] == (
        "2 of 4 mismatches in `cust_name` are truncations "
        "(avg 9.0→8.0 chars) — target column may be too narrow."
    )


def test_category_sentence_string_truncation_missing_length_stats():
    """A STRING_TRUNCATION category with no matching column_stats entry
    (e.g. caller passed {}) must say so plainly rather than crash or
    silently omit the parenthetical as if it meant something different."""
    summaries = [
        MismatchCategorySummary(
            category="STRING_TRUNCATION", count=1, pct=100.0,
            top_table="customers", top_column="cust_name",
        ),
    ]
    sentences = build_category_sentences(summaries, {})

    assert "length detail unavailable" in sentences[0]


def test_category_sentence_case_difference():
    summaries = [
        MismatchCategorySummary(
            category="CASE_DIFFERENCE", count=3, pct=100.0,
            top_table="t", top_column="c",
        ),
    ]
    assert build_category_sentences(summaries, {}) == [
        "3 mismatches are case-only — consider UPPER/LOWER normalization "
        "in the transform layer."
    ]


def test_category_sentence_whitespace_diff():
    summaries = [
        MismatchCategorySummary(
            category="WHITESPACE_DIFF", count=4, pct=100.0,
            top_table="t", top_column="c",
        ),
    ]
    assert build_category_sentences(summaries, {}) == [
        "4 mismatches differ only by whitespace — check TRIM logic in "
        "the pipeline."
    ]


def test_category_sentence_precision_loss_has_no_fabricated_numbers():
    """Must never invent a decimal-place count or a DECIMAL(p,s) guess -
    mismatch_classifier has no schema/type information to base one on."""
    summaries = [
        MismatchCategorySummary(
            category="PRECISION_LOSS", count=7, pct=100.0,
            top_table="t", top_column="amount",
        ),
    ]
    sentence = build_category_sentences(summaries, {})[0]

    assert sentence == (
        "7 numeric mismatches lose precision after conversion — verify "
        "the column's decimal/numeric type definition matches on both "
        "sides."
    )
    # No invented digits beyond the real count (7).
    assert not any(ch.isdigit() and ch != "7" for ch in sentence)


def test_category_sentence_formatting_diff():
    summaries = [
        MismatchCategorySummary(
            category="FORMATTING_DIFF", count=6, pct=100.0,
            top_table="t", top_column="order_date",
        ),
    ]
    assert build_category_sentences(summaries, {}) == [
        "6 mismatches are format-only (same value, different "
        "representation) — align date/number format in target schema."
    ]


def test_category_sentence_value_mismatch():
    summaries = [
        MismatchCategorySummary(
            category="VALUE_MISMATCH", count=9, pct=100.0,
            top_table="t", top_column="notes",
        ),
    ]
    assert build_category_sentences(summaries, {}) == [
        "9 mismatches have no detectable pattern — manual review "
        "recommended for `notes`."
    ]


def test_category_sentences_one_per_category_in_given_order():
    summaries = [
        MismatchCategorySummary(category="CASE_DIFFERENCE", count=3, pct=60.0, top_table="t", top_column="a"),
        MismatchCategorySummary(category="NULL_MISMATCH", count=2, pct=40.0, top_table="t", top_column="b"),
    ]
    sentences = build_category_sentences(summaries, {})

    assert len(sentences) == 2
    assert "case-only" in sentences[0]
    assert "nulls mismatched" in sentences[1]


def test_category_sentences_empty_for_empty_summaries():
    assert build_category_sentences([], {}) == []


# ---------------------------------------------------------------------------
# Phase 4: "Mismatch Categories" sheet - additive only. Data Mismatches and
# Suggestions must render byte-identically to before this feature, and the
# new sheet must show up in the right place, with the right two sections,
# gated the same way as the other row-level sheets.
# ---------------------------------------------------------------------------
def _response_with_mismatches(*details: RowMismatchDetail) -> CatalogValidationResponse:
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        exists_in_source=True, exists_in_target=True,
        data=DataValidationResult(
            mode="FULL", status=ValidationStatus.FAIL,
            sample_changed_detail=list(details),
        ),
    )
    schema = SchemaValidationResult(
        schema_name="bronze", status=ValidationStatus.FAIL, tables=[table],
    )
    return _response(schema)


def test_mismatch_categories_sheet_present_and_positioned_before_suggestions(tmp_path):
    detail = RowMismatchDetail(
        schema_name="bronze", table="customers",
        primary_key={"id": 1}, mismatch_column="status",
        source_value="ACTIVE", target_value="active",
    )
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(_response_with_mismatches(detail), str(output_path))

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    assert "Mismatch Categories" in wb.sheetnames
    # Right after Row Hash Mismatches, before Suggestions.
    names = wb.sheetnames
    assert names.index("Row Hash Mismatches") < names.index("Mismatch Categories") < names.index("Suggestions")


def test_mismatch_categories_sheet_gated_by_row_validation(tmp_path):
    detail = RowMismatchDetail(
        schema_name="bronze", table="customers",
        primary_key={"id": 1}, mismatch_column="status",
        source_value="ACTIVE", target_value="active",
    )
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(
        _response_with_mismatches(detail), str(output_path),
        enabled_validations={"catalog", "schema", "column"},
    )

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    assert "Mismatch Categories" not in wb.sheetnames
    # The sheets it's derived from must be gone too, for the same reason.
    assert "Data Mismatches" not in wb.sheetnames


def test_mismatch_categories_sheet_summary_table_section(tmp_path):
    detail_a = RowMismatchDetail(
        schema_name="bronze", table="customers",
        primary_key={"id": 1}, mismatch_column="status",
        source_value="ACTIVE", target_value="active",
    )
    detail_b = RowMismatchDetail(
        schema_name="bronze", table="customers",
        primary_key={"id": 2}, mismatch_column="notes",
        source_value=None, target_value="x",
    )
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(_response_with_mismatches(detail_a, detail_b), str(output_path))

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb["Mismatch Categories"]

    assert [c.value for c in ws[1]] == ["Category", "Count", "% of Total", "Top Table", "Top Column"]
    body_values = {ws.cell(row=r, column=1).value for r in (2, 3)}
    assert body_values == {"CASE_DIFFERENCE", "NULL_MISMATCH"}
    # Count/% columns actually populated, not blank.
    counts = {ws.cell(row=r, column=2).value for r in (2, 3)}
    assert counts == {1, 1}


def test_mismatch_categories_sheet_category_column_colored(tmp_path):
    detail = RowMismatchDetail(
        schema_name="bronze", table="customers",
        primary_key={"id": 1}, mismatch_column="status",
        source_value="ACTIVE", target_value="active",  # CASE_DIFFERENCE
    )
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(_response_with_mismatches(detail), str(output_path))

    from openpyxl import load_workbook
    from table_validator.reports.excel_report import STATUS_FILLS
    wb = load_workbook(output_path)
    ws = wb["Mismatch Categories"]

    category_cell = ws.cell(row=2, column=1)
    assert category_cell.value == "CASE_DIFFERENCE"
    assert category_cell.fill.fgColor.rgb == STATUS_FILLS[ValidationStatus.PASS].fgColor.rgb


def test_mismatch_categories_sheet_insight_sentences_section(tmp_path):
    detail = RowMismatchDetail(
        schema_name="bronze", table="customers",
        primary_key={"id": 1}, mismatch_column="status",
        source_value="ACTIVE", target_value="active",
    )
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(_response_with_mismatches(detail), str(output_path))

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb["Mismatch Categories"]

    all_values = [
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert "Insight Sentences" in all_values
    assert any("case-only" in str(v) for v in all_values)


def test_mismatch_categories_sheet_empty_when_no_mismatches(tmp_path):
    """No row-level mismatches at all (e.g. every table PASSed) must still
    produce a (header-only) sheet, not crash or silently omit it, when
    row-level validation is otherwise enabled."""
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.PASS,
        data=DataValidationResult(mode="STATISTICS", status=ValidationStatus.PASS),
    )
    schema = SchemaValidationResult(schema_name="bronze", status=ValidationStatus.PASS, tables=[table])
    output_path = tmp_path / "report.xlsx"

    generate_excel_report(_response(schema), str(output_path))

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb["Mismatch Categories"]
    assert [c.value for c in ws[1]] == CATEGORY_SUMMARY_HEADERS
    assert ws.max_row == 1


def test_data_mismatches_sheet_unaffected_by_new_category_sheet(tmp_path):
    """Regression guard: the new Mismatch Categories sheet must be purely
    additive - Data Mismatches' own rows/headers must render identically
    to before this feature existed."""
    detail = RowMismatchDetail(
        schema_name="bronze", table="customers",
        primary_key={"id": 1}, mismatch_column="status",
        source_value="ACTIVE", target_value="active",
    )
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(_response_with_mismatches(detail), str(output_path))

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb["Data Mismatches"]
    assert [c.value for c in ws[1]] == MISMATCH_HEADERS
    assert ws.cell(row=2, column=4).value == "status"
    assert ws.cell(row=2, column=5).value == "ACTIVE"
    assert ws.cell(row=2, column=6).value == "active"


def test_suggestions_sheet_unaffected_by_new_category_sheet(tmp_path):
    """Regression guard: Suggestions' own row-building/formatting must be
    untouched by the new sheet's presence."""
    table = TableValidationResult(
        schema_name="bronze", table="customers", status=ValidationStatus.FAIL,
        missing_columns=["ssn"],
    )
    schema = SchemaValidationResult(schema_name="bronze", status=ValidationStatus.FAIL, tables=[table])
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(_response(schema), str(output_path))

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb["Suggestions"]
    assert [c.value for c in ws[1]] == SUGGESTION_HEADERS
    # This fixture also triggers the pre-existing "Summary" row (row data
    # matches, failure is schema-only) ahead of the specific issue row -
    # unrelated to the new sheet, just this fixture's own shape.
    assert ws.cell(row=3, column=4).value == "Missing Column"


# ---------------------------------------------------------------------------
# Integration pass: --skip-category-summary / skip_category_summary must
# recover the exact pre-categorization report, byte-for-byte on every sheet
# other than Mismatch Categories itself (which is the one sheet it omits).
# ---------------------------------------------------------------------------
def _sheet_snapshot(ws) -> list:
    """All non-empty cell values, in reading order, plus each cell's fill
    color where set - a simple byte-for-byte-equivalent fingerprint of a
    sheet's rendered content and status-coloring, without depending on
    openpyxl's exact internal XML byte layout (which can differ between
    two separately-saved .xlsx files with identical visible content, e.g.
    in embedded timestamps/zip metadata)."""
    snapshot = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
            snapshot.append((cell.coordinate, cell.value, fill))
    return snapshot


def test_skip_category_summary_omits_only_that_sheet(tmp_path):
    detail = RowMismatchDetail(
        schema_name="bronze", table="customers",
        primary_key={"id": 1}, mismatch_column="status",
        source_value="ACTIVE", target_value="active",
    )
    response = _response_with_mismatches(detail)

    with_categories_path = tmp_path / "with_categories.xlsx"
    without_categories_path = tmp_path / "without_categories.xlsx"
    generate_excel_report(response, str(with_categories_path))
    generate_excel_report(response, str(without_categories_path), skip_category_summary=True)

    from openpyxl import load_workbook
    wb_with = load_workbook(with_categories_path)
    wb_without = load_workbook(without_categories_path)

    assert "Mismatch Categories" in wb_with.sheetnames
    assert "Mismatch Categories" not in wb_without.sheetnames

    # Every OTHER sheet must render identically regardless of the flag.
    other_sheets = [name for name in wb_with.sheetnames if name != "Mismatch Categories"]
    assert other_sheets == wb_without.sheetnames
    for sheet_name in other_sheets:
        assert _sheet_snapshot(wb_with[sheet_name]) == _sheet_snapshot(wb_without[sheet_name]), (
            f"Sheet '{sheet_name}' differs with skip_category_summary=True"
        )


def test_skip_category_summary_data_mismatches_sheet_matches_pre_feature_shape(tmp_path):
    """Directly pins Data Mismatches' exact rendered content with the flag
    on, independent of the comparison-based test above - the specific
    regression this integration pass is meant to prevent."""
    detail = RowMismatchDetail(
        schema_name="bronze", table="customers",
        primary_key={"id": 1}, mismatch_column="status",
        source_value="ACTIVE", target_value="active", verified=True,
    )
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(
        _response_with_mismatches(detail), str(output_path), skip_category_summary=True,
    )

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb["Data Mismatches"]

    assert [c.value for c in ws[1]] == MISMATCH_HEADERS
    row = [ws.cell(row=2, column=c).value for c in range(1, len(MISMATCH_HEADERS) + 1)]
    assert row == ["bronze", "customers", "id=1", "status", "ACTIVE", "active", None, None, "Yes"]


def test_skip_category_summary_default_is_false_and_includes_sheet(tmp_path):
    """The flag must be strictly opt-in - a caller that never mentions it
    (matching every pre-existing caller of generate_excel_report) keeps
    getting the Mismatch Categories sheet."""
    detail = RowMismatchDetail(
        schema_name="bronze", table="customers",
        primary_key={"id": 1}, mismatch_column="status",
        source_value="ACTIVE", target_value="active",
    )
    output_path = tmp_path / "report.xlsx"
    generate_excel_report(_response_with_mismatches(detail), str(output_path))

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    assert "Mismatch Categories" in wb.sheetnames
