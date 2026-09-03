"""
Report Generator (CSV + Excel)

Consumes a CatalogValidationResponse (produced by
validators.catalog_validator.CatalogValidator) and renders it as either a flat .csv
file (Table Validation data only) or a formatted, multi-sheet .xlsx
workbook (Summary / Table Validation / Column Validation / Data
Mismatches / Row Hash Mismatches / Mismatch Categories / Suggestions).

Deliberately kept out of comparison_engine.py per the project spec:
the validator returns clean structured data, and this module is the
only place that knows about report formatting (csv / openpyxl).

Mismatch Categories sheet (root-cause classification):
Every RowMismatchDetail already collected under Data Mismatches (the
result of the validator's own row-hash/column-diff logic - this module
adds no new comparisons, no new Databricks queries) is additionally
classified into one of validators.mismatch_classifier's fixed labels -
NULL_MISMATCH, STRING_TRUNCATION, CASE_DIFFERENCE, WHITESPACE_DIFF,
PRECISION_LOSS, FORMATTING_DIFF, VALUE_MISMATCH - purely from the two
already-fetched source/target values (see mismatch_classifier.py's own
docstring for the exact rule order). generate_excel_report() runs this
classification + aggregation + sentence-generation pipeline
(_classify_all_mismatches -> _build_mismatch_category_summary ->
build_category_sentences) internally, right before writing the Mismatch
Categories sheet, whenever row-level validation is enabled - it never
touches how Data Mismatches/Row Hash Mismatches/any other sheet is
built, and RowMismatchDetail.mismatch_category is the only field it
writes back onto the response (see models.py). Pass
skip_category_summary=True (or the CLI's --skip-category-summary) to
omit just this sheet and get the exact pre-categorization report back,
byte-for-byte, on every other sheet.
"""

from __future__ import annotations

import csv
import logging
from collections import Counter, defaultdict
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from table_validator.models import (
    CatalogValidationResponse,
    RowMismatchDetail,
    ValidationStatus,
)
from table_validator.validators.mismatch_classifier import (
    CASE_DIFFERENCE,
    FORMATTING_DIFF,
    NULL_MISMATCH,
    PRECISION_LOSS,
    STRING_TRUNCATION,
    VALUE_MISMATCH,
    WHITESPACE_DIFF,
    classify_mismatch,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Shared column definitions
# ---------------------------------------------------------------------------
TABLE_HEADERS = [
    "Source Schema",
    "Source Table",
    "Target Schema",
    "Target Table",
    "Overall Status",
    "Schema Match",
    "Column Order",
    "Row Count (Src)",
    "Row Count (Tgt)",
    "Row Count Diff",
    "Data Types",
    "Nullable",
    "Null Counts",
    "Distinct Counts",
    "Min/Max",
    "Data Match",
    "Mismatch Count",
    "Mismatch %",
    "Row Hash Mismatch Count",
    "Row Hash Mismatch %",
    "Tier Reached",
    "Partition",
    "Validation Timestamp",
    "Duration",
]

# 1-based column indices within TABLE_HEADERS holding a PASS/FAIL/ERROR/SKIPPED value.
_TABLE_STATUS_COLUMNS = {5, 6, 7, 11, 12, 13, 14, 15, 16}

# Which ValidationType (config/schema.py) owns each 0-based TABLE_HEADERS
# column, for hiding columns whose validation type wasn't selected to run
# (see _filter_table_columns). None = always shown regardless of
# enabled_validations. "Schema Match"/"Column Order" measure column-name/
# order agreement (COLUMN, despite the "Schema Match" label being a
# pre-existing naming quirk from before source types other than
# Databricks existed) - not schema/table existence, which has no
# dedicated column here (it's surfaced as separate MISSING_FROM_TARGET/
# EXTRA_IN_TARGET rows instead, unaffected by column filtering).
_TABLE_COLUMN_OWNERS = [
    None, None, None, None, None,  # Source/Target Schema/Table, Overall Status
    "column",   # Schema Match
    "column",   # Column Order
    "row", "row", "row",  # Row Count (Src/Tgt/Diff)
    "column",   # Data Types
    "column",   # Nullable
    "column",   # Null Counts
    "column",   # Distinct Counts
    "column",   # Min/Max
    "row",      # Data Match
    "row", "row",  # Mismatch Count/%
    "row", "row",  # Row Hash Mismatch Count/%
    "row",      # Tier Reached
    "row",      # Partition
    None, None,  # Validation Timestamp, Duration
]

COLUMN_HEADERS = [
    "Source Schema", "Source Table", "Column", "Status",
    "Source Type", "Target Type", "Type Status",
    "Source Nullable", "Target Nullable", "Nullable Status",
    "Source Nulls", "Target Nulls", "Null Status",
    "Source Distinct", "Target Distinct", "Distinct Status",
    "Source Min", "Source Max", "Target Min", "Target Max", "Min/Max Status",
    "Error",
]
_COLUMN_STATUS_COLUMNS = {4, 7, 10, 13, 16, 21}

MISMATCH_HEADERS = [
    "Source Schema", "Source Table", "Primary Key",
    "Mismatch Column", "Source Value", "Target Value",
    "Row Hash (Source)", "Row Hash (Target)", "Verified",
]

ROW_HASH_HEADERS = [
    "Row #", "Source Schema", "Source Table",
    "Primary Key", "Source Hash", "Target Hash", "Mismatch Status",
    "Partition Bucket",
]
_ROW_HASH_STATUS_COLUMN = 7

SUGGESTION_HEADERS = [
    "Source Schema", "Source Table", "Column", "Issue Type", "Suggestion",
]

CATEGORY_SUMMARY_HEADERS = [
    "Category", "Count", "% of Total", "Top Table", "Top Column",
]

SUMMARY_METRIC_LABELS = [
    "Total Tables", "Passed Tables", "Failed Tables",
    "Error Tables", "Skipped Tables", "Pass Percentage",
]

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="1F4E78")
LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True)
VALUE_FONT = Font(name=FONT_NAME, size=10)
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

STATUS_FILLS = {
    ValidationStatus.PASS: PatternFill("solid", fgColor="C6EFCE"),
    ValidationStatus.FAIL: PatternFill("solid", fgColor="FFC7CE"),
    ValidationStatus.ERROR: PatternFill("solid", fgColor="FFD8A8"),
    ValidationStatus.SKIPPED: PatternFill("solid", fgColor="E7E6E6"),
}
STATUS_FONTS = {
    ValidationStatus.PASS: Font(name=FONT_NAME, size=10, bold=True, color="006100"),
    ValidationStatus.FAIL: Font(name=FONT_NAME, size=10, bold=True, color="9C0006"),
    ValidationStatus.ERROR: Font(name=FONT_NAME, size=10, bold=True, color="974706"),
    ValidationStatus.SKIPPED: Font(name=FONT_NAME, size=10, bold=True, color="666666"),
}

# Sort rank so FAIL/ERROR surface above PASS/SKIPPED within a schema.
_STATUS_SORT_RANK = {
    "FAIL": 0,
    "ERROR": 0,
    "MISSING_FROM_TARGET": 0,
    "EXTRA_IN_TARGET": 0,
    "PASS": 1,
    "SKIPPED": 1,
}

# Row-hash mismatch statuses aren't ValidationStatus members - map them onto
# the same PASS/FAIL/ERROR/SKIPPED fill+font conventions for the Row Hash
# Mismatches sheet (all three outcomes here are a real difference -> FAIL).
_ROW_HASH_STATUS_FILL_MAP = {
    "MISMATCH": ValidationStatus.FAIL,
    "MISSING_IN_TARGET": ValidationStatus.FAIL,
    "MISSING_IN_SOURCE": ValidationStatus.FAIL,
    "DUPLICATE_KEY": ValidationStatus.FAIL,
}

# Conditional Category-column color per mismatch_classifier label, for the
# Mismatch Categories sheet - reuses the exact same fill/font pairs as
# STATUS_FILLS/STATUS_FONTS above rather than inventing new colors, so the
# whole workbook stays visually consistent (red=FAIL-like/severe, orange=
# ERROR-like/needs attention, green=PASS-like/benign, gray=neutral/
# uncategorized-pattern). Severity here is a judgment call about how
# likely each category is to indicate a real data/schema problem, not a
# property mismatch_classifier itself computes:
#   - NULL_MISMATCH, VALUE_MISMATCH: most likely a genuine data problem
#     (lost data, unexplained change) -> red (FAIL).
#   - STRING_TRUNCATION, PRECISION_LOSS: a real schema/type sizing issue,
#     but a well-understood, mechanically fixable one -> orange (ERROR).
#   - CASE_DIFFERENCE, WHITESPACE_DIFF, FORMATTING_DIFF: same underlying
#     value, cosmetic/representational only -> green (PASS-like/benign).
_CATEGORY_STATUS_MAP = {
    NULL_MISMATCH: ValidationStatus.FAIL,
    VALUE_MISMATCH: ValidationStatus.FAIL,
    STRING_TRUNCATION: ValidationStatus.ERROR,
    PRECISION_LOSS: ValidationStatus.ERROR,
    CASE_DIFFERENCE: ValidationStatus.PASS,
    WHITESPACE_DIFF: ValidationStatus.PASS,
    FORMATTING_DIFF: ValidationStatus.PASS,
}


def _status_value(status: Optional[ValidationStatus]) -> str:
    if status is None:
        return ""
    return status.value if isinstance(status, ValidationStatus) else str(status)


def _mismatch_count(table) -> Optional[int]:
    if table.data is None:
        return None
    counts = [table.data.source_only_rows, table.data.target_only_rows, table.data.changed_rows]
    if not all(c is None for c in counts):
        return sum(c or 0 for c in counts)
    # The tiered fail-fast funnel (Databricks-to-Databricks) never
    # populates source_only_rows/target_only_rows/changed_rows - those
    # only ever came from the legacy FULL-mode EXCEPT/hash-join path.
    # Fall back to the row-hash comparison's own mismatch count, which is
    # the real per-row mismatch figure for every table run through the
    # tiered pipeline.
    if table.data.row_hash_mismatch_count:
        return table.data.row_hash_mismatch_count
    return None


def _partition_summary(table) -> str:
    if table.partitioned:
        return (
            f"{table.partition_column} "
            f"({table.partition_buckets_culprit}/{table.partition_buckets_total} buckets differed)"
        )
    if table.partition_skip_reason:
        return f"not partitioned ({table.partition_skip_reason})"
    return ""


def _mismatch_pct(table, mismatch_count: Optional[int]) -> str:
    if mismatch_count is None:
        return ""
    if table.data is not None and (
        table.data.source_only_rows is None
        and table.data.target_only_rows is None
        and table.data.changed_rows is None
    ):
        # Tiered-pipeline fallback (see _mismatch_count): reuse the
        # row-hash comparison's own percentage, which is computed over
        # the union of keys actually compared on either side - more
        # accurate than dividing by row_count_source alone when keys are
        # missing/extra on one side.
        return f"{table.data.row_hash_mismatch_percentage:.2f}%"
    total = table.row_count_source or 0
    if total <= 0:
        return ""
    return f"{(mismatch_count / total) * 100:.2f}%"


def _filter_table_columns(
    headers: List[str],
    rows: List[List[Any]],
    enabled_validations: Optional[set] = None,
) -> Tuple[List[str], List[List[Any]], set]:
    """
    Drop TABLE_HEADERS columns (and the matching cell in every row) whose
    owning validation type isn't in enabled_validations - e.g. hide "Data
    Types"/"Null Counts"/etc. entirely when "column" wasn't selected to
    run, rather than showing a column of meaningless SKIPPED values.

    enabled_validations=None means "no filtering" (show everything),
    matching the CSV/legacy behavior and any caller that doesn't know
    about per-run validation selection.

    Returns (filtered_headers, filtered_rows, status_columns) where
    status_columns is the 1-based set of surviving columns that hold a
    PASS/FAIL/ERROR/SKIPPED value, recomputed for the new column
    positions (_TABLE_STATUS_COLUMNS' original indices no longer apply
    once columns are dropped).
    """
    if enabled_validations is None:
        return headers, rows, _TABLE_STATUS_COLUMNS

    keep_indices = [
        i for i, owner in enumerate(_TABLE_COLUMN_OWNERS)
        if owner is None or owner in enabled_validations
    ]

    filtered_headers = [headers[i] for i in keep_indices]
    filtered_rows = [[row[i] for i in keep_indices] for row in rows]
    status_columns = {
        new_idx + 1
        for new_idx, old_idx in enumerate(keep_indices)
        if (old_idx + 1) in _TABLE_STATUS_COLUMNS
    }

    return filtered_headers, filtered_rows, status_columns


# ---------------------------------------------------------------------------
# Row builders (shared by CSV + every Excel sheet)
# ---------------------------------------------------------------------------
def _build_table_rows(result: CatalogValidationResponse) -> List[List[Any]]:
    timestamp = result.validation_timestamp or ""
    duration = result.execution_time_seconds

    rows: List[List[Any]] = []
    for schema in result.schemas:
        for table in schema.tables:
            data_status = table.data.status if table.data else None
            mismatch_count = _mismatch_count(table)

            row_hash_count = table.data.row_hash_mismatch_count if table.data else 0
            row_hash_pct = (
                f"{table.data.row_hash_mismatch_percentage:.2f}%" if table.data else ""
            )

            rows.append(
                [
                    schema.schema_name,
                    table.table,
                    schema.schema_name,
                    table.table,
                    _status_value(table.status),
                    _status_value(table.columns_status),
                    _status_value(table.column_order_status),
                    table.row_count_source,
                    table.row_count_target,
                    table.row_count_difference,
                    _status_value(table.data_types_status),
                    _status_value(table.nullable_status),
                    _status_value(table.null_counts_status),
                    _status_value(table.distinct_counts_status),
                    _status_value(table.min_max_status),
                    _status_value(data_status),
                    mismatch_count if mismatch_count is not None else "",
                    _mismatch_pct(table, mismatch_count),
                    row_hash_count,
                    row_hash_pct,
                    table.tier_reached.name,
                    _partition_summary(table),
                    timestamp,
                    duration,
                ]
            )

        for missing_table in schema.missing_tables:
            rows.append(
                [
                    schema.schema_name, missing_table, schema.schema_name, missing_table,
                    "MISSING_FROM_TARGET",
                ]
                + [""] * 17
                + [timestamp, duration]
            )

        for extra_table in schema.extra_tables:
            rows.append(
                [
                    schema.schema_name, extra_table, schema.schema_name, extra_table,
                    "EXTRA_IN_TARGET",
                ]
                + [""] * 17
                + [timestamp, duration]
            )

    # Sort: Source Schema asc, then FAIL/ERROR before PASS/SKIPPED, then Source Table asc.
    rows.sort(key=lambda r: (r[0], _STATUS_SORT_RANK.get(r[4], 1), r[1]))
    return rows


def _build_column_rows(result: CatalogValidationResponse) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for schema in result.schemas:
        for table in schema.tables:
            for col in table.columns:
                rows.append(
                    [
                        schema.schema_name, table.table, col.column, _status_value(col.status),
                        col.source_data_type, col.target_data_type, _status_value(col.data_type_status),
                        col.source_nullable, col.target_nullable, _status_value(col.nullable_status),
                        col.source_null_count, col.target_null_count, _status_value(col.null_count_status),
                        col.source_distinct_count, col.target_distinct_count, _status_value(col.distinct_count_status),
                        col.source_min, col.source_max, col.target_min, col.target_max,
                        _status_value(col.min_max_status),
                        col.error or "",
                    ]
                )

            for missing_col in table.missing_columns:
                rows.append(
                    [schema.schema_name, table.table, missing_col, "MISSING_FROM_TARGET"]
                    + [""] * 18
                )
            for extra_col in table.extra_columns:
                rows.append(
                    [schema.schema_name, table.table, extra_col, "EXTRA_IN_TARGET"]
                    + [""] * 18
                )

    rows.sort(key=lambda r: (r[0], r[1], r[2]))
    return rows


# ---------------------------------------------------------------------------
# Mismatch categorization (phase 2): classify every RowMismatchDetail
# already collected by the validator, then roll the results up into a
# per-category summary. Pure in-memory computation over data the validator
# already produced - no new Databricks queries, no Excel writing here (the
# summary is consumed by a later phase's sheet-builder).
# ---------------------------------------------------------------------------
@dataclass
class MismatchCategorySummary:
    """One row of the eventual category-summary view: how often a given
    mismatch_classifier category occurred, what share of all mismatches
    that represents, and which table/column it showed up in most - a
    quick pointer for "where do I look first for this kind of issue"
    without having to scan the full Data Mismatches sheet."""

    category: str
    count: int
    pct: float
    top_table: str
    top_column: str


def _classify_all_mismatches(result: CatalogValidationResponse) -> List[RowMismatchDetail]:
    """
    Classify every RowMismatchDetail in `result` in place, setting its
    mismatch_category field via classify_mismatch(), and return the flat
    list of details classified (same traversal order as
    _build_mismatch_rows, so both stay consistent).

    Mutates the response in place (rather than building a parallel
    structure) so mismatch_category travels with each RowMismatchDetail
    wherever it goes next - the Data Mismatches sheet, a future JSON
    response, etc. - matching the field's own docstring in models.py.
    Idempotent: re-running this on an already-classified response just
    recomputes the same labels.
    """
    all_details: List[RowMismatchDetail] = []
    for schema in result.schemas:
        for table in schema.tables:
            if table.data is None:
                continue
            for detail in table.data.sample_changed_detail:
                detail.mismatch_category = classify_mismatch(
                    detail.source_value, detail.target_value
                )
                all_details.append(detail)
    return all_details


def _build_mismatch_category_summary(
    result: CatalogValidationResponse,
) -> List[MismatchCategorySummary]:
    """
    Classify every mismatch (via _classify_all_mismatches) and roll the
    results up into one MismatchCategorySummary per category that
    actually occurred at least once - count, percentage of the total
    mismatch count, and the table/column each category shows up in most
    often (its single biggest concentration, not an exhaustive list -
    the Data Mismatches sheet already has full per-row detail).

    Percentages are computed over the total number of classified
    mismatches across every table/schema in `result`, not per-table -
    "42% of all mismatches are VALUE_MISMATCH" is the intended reading.
    Returns [] when there are no mismatches to classify at all (e.g.
    every table PASSed, or row-level comparison never ran).
    """
    all_details = _classify_all_mismatches(result)
    total = len(all_details)
    if total == 0:
        return []

    category_counts: Counter = Counter()
    # category -> table -> count, and category -> column -> count, kept
    # separately since "top table" and "top column" for a category are
    # independent questions (the biggest table by mismatch count for a
    # category need not share its name with the biggest column).
    category_table_counts: Dict[str, Counter] = defaultdict(Counter)
    category_column_counts: Dict[str, Counter] = defaultdict(Counter)

    for detail in all_details:
        category = detail.mismatch_category
        category_counts[category] += 1
        category_table_counts[category][detail.table] += 1
        # source_mismatch_column is only set when column_map actually
        # renamed this column - mismatch_column (the canonical/target
        # name) is what every other sheet groups by, so use it here too
        # for consistency.
        category_column_counts[category][detail.mismatch_column] += 1

    summaries: List[MismatchCategorySummary] = []
    for category, count in category_counts.items():
        top_table, _ = category_table_counts[category].most_common(1)[0]
        top_column, _ = category_column_counts[category].most_common(1)[0]
        summaries.append(
            MismatchCategorySummary(
                category=category,
                count=count,
                pct=round(count / total * 100, 2),
                top_table=top_table,
                top_column=top_column,
            )
        )

    # Most common category first - the natural "look at this one first"
    # reading order; ties broken alphabetically for deterministic output.
    summaries.sort(key=lambda s: (-s.count, s.category))
    return summaries


# ---------------------------------------------------------------------------
# Phase 3: plain-English sentences per category, for a future Suggestions-
# style sheet. Pure string formatting over Phase 2's summaries plus a
# small per-column length stat (needed only for STRING_TRUNCATION's
# avg-chars wording) - no Excel writing here yet, no new queries.
# ---------------------------------------------------------------------------
@dataclass
class ColumnLenStats:
    """Average source/target string length for one column, across every
    mismatch on it that was actually classified as STRING_TRUNCATION -
    the only category whose sentence needs raw value lengths rather than
    just a count. Built from the same classified RowMismatchDetail list
    Phase 2's _classify_all_mismatches() already produces."""

    avg_source_len: float
    avg_target_len: float


def build_column_len_stats(
    details: List[RowMismatchDetail],
) -> Dict[str, ColumnLenStats]:
    """
    Compute average source/target value length per column, over only the
    details already classified as STRING_TRUNCATION (mismatch_category
    must be set - i.e. `details` should come from
    _classify_all_mismatches(), not a raw, unclassified list). Columns
    with no STRING_TRUNCATION mismatches are simply absent from the
    result - there's nothing meaningful to average.

    Grouped by mismatch_column (the canonical/target name), matching
    _build_mismatch_category_summary's own convention for a renamed
    column. Non-string values are skipped (len() on a coerced str() would
    silently fabricate a "length" for a number/None, which STRING_
    TRUNCATION's classification already restricts to actual strings -
    this mirrors that same restriction here rather than reinterpreting
    the rule).
    """
    source_lens: Dict[str, List[int]] = defaultdict(list)
    target_lens: Dict[str, List[int]] = defaultdict(list)

    for detail in details:
        if detail.mismatch_category != STRING_TRUNCATION:
            continue
        if not isinstance(detail.source_value, str) or not isinstance(detail.target_value, str):
            continue
        source_lens[detail.mismatch_column].append(len(detail.source_value))
        target_lens[detail.mismatch_column].append(len(detail.target_value))

    return {
        column: ColumnLenStats(
            avg_source_len=sum(source_lens[column]) / len(source_lens[column]),
            avg_target_len=sum(target_lens[column]) / len(target_lens[column]),
        )
        for column in source_lens
    }


def build_category_sentences(
    summaries: List[MismatchCategorySummary],
    column_stats: Dict[str, ColumnLenStats],
) -> List[str]:
    """
    One plain-English sentence per category present in `summaries` -
    honest and specific, built only from numbers Phase 1/2 actually
    computed (counts, percentages, top table/column, and - for
    STRING_TRUNCATION only - the avg source/target length in
    column_stats). Deliberately no fabricated confidence scores, decimal-
    place counts, or DECIMAL(p,s) guesses: mismatch_classifier never sees
    column type/schema information (by Phase 1's own design), so nothing
    here invents numbers it doesn't have - PRECISION_LOSS's sentence
    stays qualitative rather than naming a decimal count it can't derive.

    STRING_TRUNCATION's "{count} of {total}" wording uses the grand total
    across every category, computed here as sum(s.count for s in
    summaries) - exact as long as `summaries` is the complete list from
    _build_mismatch_category_summary (true for every real caller), unlike
    reconstructing it from MismatchCategorySummary's already-rounded pct
    (which would drift, and divide by zero at pct == 0).

    Order follows `summaries` as given (Phase 2 already sorts most-common
    category first) - this function doesn't re-sort. A category with no
    template below (there shouldn't be one, since mismatch_classifier's
    label set is fixed and total) is skipped rather than raising, so a
    future new category doesn't crash report generation before its own
    sentence template is added.
    """
    total = sum(s.count for s in summaries)
    sentences: List[str] = []

    for summary in summaries:
        if summary.category == NULL_MISMATCH:
            sentences.append(
                f"{summary.count} nulls mismatched in {summary.top_column} — "
                f"check nullable constraint differences between source and "
                f"target."
            )

        elif summary.category == STRING_TRUNCATION:
            stats = column_stats.get(summary.top_column)
            if stats is not None:
                src_len = round(stats.avg_source_len, 1)
                tgt_len = round(stats.avg_target_len, 1)
                length_note = f" (avg {src_len}→{tgt_len} chars)"
            else:
                # No length stats available for this column (e.g. caller
                # passed an empty/mismatched column_stats) - say so
                # rather than silently omitting the parenthetical or
                # inventing a length.
                length_note = " (length detail unavailable)"
            sentences.append(
                f"{summary.count} of {total} mismatches in "
                f"`{summary.top_column}` are truncations{length_note} — "
                f"target column may be too narrow."
            )

        elif summary.category == CASE_DIFFERENCE:
            sentences.append(
                f"{summary.count} mismatches are case-only — consider "
                f"UPPER/LOWER normalization in the transform layer."
            )

        elif summary.category == WHITESPACE_DIFF:
            sentences.append(
                f"{summary.count} mismatches differ only by whitespace — "
                f"check TRIM logic in the pipeline."
            )

        elif summary.category == PRECISION_LOSS:
            sentences.append(
                f"{summary.count} numeric mismatches lose precision after "
                f"conversion — verify the column's decimal/numeric type "
                f"definition matches on both sides."
            )

        elif summary.category == FORMATTING_DIFF:
            sentences.append(
                f"{summary.count} mismatches are format-only (same value, "
                f"different representation) — align date/number format in "
                f"target schema."
            )

        elif summary.category == VALUE_MISMATCH:
            sentences.append(
                f"{summary.count} mismatches have no detectable pattern — "
                f"manual review recommended for `{summary.top_column}`."
            )

    return sentences


def _build_mismatch_category_sheet(wb: Workbook, result: CatalogValidationResponse) -> None:
    """
    Phase 4: renders Phase 2/3's category summary + sentences as their own
    "Mismatch Categories" sheet - purely additive (a new sheet; the Data
    Mismatches and Suggestions sheets are untouched by this function and
    built by their own separate _build_*_sheet calls).

    Two sections on one sheet, stacked vertically rather than split across
    sheets, so a reader sees the "what" (Section 1's table) right above
    the "why/what to do" (Section 2's sentences) in one place:
      Section 1 - Summary Table: one row per MismatchCategorySummary
        (Category | Count | % of Total | Top Table | Top Column), header-
        styled like every other sheet's table, with each Category cell
        colored via _CATEGORY_STATUS_MAP/STATUS_FILLS/STATUS_FONTS - the
        exact same fill/font pairs the rest of the workbook already uses
        for PASS/FAIL/ERROR, so this sheet doesn't introduce a second,
        inconsistent color language.
      Section 2 - Insight Sentences: Phase 3's plain-English sentences,
        one per row, directly below Section 1 with a blank row and its
        own sub-heading between them.

    Skips rendering entirely (still creates an empty, header-only sheet)
    when there are no mismatches to categorize - matches how e.g. Row
    Hash Mismatches renders zero data rows rather than omitting itself,
    once its owning validation type is enabled.
    """
    ws = wb.create_sheet("Mismatch Categories")

    summaries = _build_mismatch_category_summary(result)
    classified = _classify_all_mismatches(result)
    column_stats = build_column_len_stats(classified)
    sentences = build_category_sentences(summaries, column_stats)

    # --- Section 1: Summary Table -----------------------------------
    _write_header_row(ws, CATEGORY_SUMMARY_HEADERS)

    row_idx = 2
    for summary in summaries:
        values = [
            summary.category,
            summary.count,
            f"{summary.pct:.2f}%",
            summary.top_table,
            summary.top_column,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if col_idx == 1:
                # Conditional Category-column color - falls back to no
                # fill/font override for a category not in the map
                # (there shouldn't be one, since mismatch_classifier's
                # label set is fixed, but this stays defensive rather
                # than raising over a purely cosmetic mapping).
                status = _CATEGORY_STATUS_MAP.get(summary.category)
                if status in STATUS_FILLS:
                    cell.fill = STATUS_FILLS[status]
                    cell.font = STATUS_FONTS[status]
        row_idx += 1

    last_table_row = max(row_idx - 1, 1)
    _autofit(
        ws, CATEGORY_SUMMARY_HEADERS,
        [
            [s.category, s.count, f"{s.pct:.2f}%", s.top_table, s.top_column]
            for s in summaries
        ],
    )
    _enable_filter(ws, len(CATEGORY_SUMMARY_HEADERS), last_table_row)

    # --- Section 2: Insight Sentences -------------------------------
    # Two blank rows separate the auto-filtered table above from the
    # free-text sentences below, so the filter's own dropdown-header row
    # is never mistaken for part of this second section. Skipped
    # entirely (no heading, no rows) when there are no mismatches to
    # categorize at all - an "Insight Sentences" heading with nothing
    # under it would be a confusing empty section, not a useful one.
    if sentences:
        sentences_header_row = last_table_row + 3
        heading_cell = ws.cell(row=sentences_header_row, column=1, value="Insight Sentences")
        heading_cell.font = LABEL_FONT

        sentence_row = sentences_header_row + 1
        for sentence in sentences:
            cell = ws.cell(row=sentence_row, column=1, value=sentence)
            cell.font = VALUE_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.merge_cells(
                start_row=sentence_row, start_column=1,
                end_row=sentence_row, end_column=len(CATEGORY_SUMMARY_HEADERS),
            )
            sentence_row += 1

    ws.freeze_panes = ws.cell(row=2, column=1)


def _build_mismatch_rows(result: CatalogValidationResponse) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for schema in result.schemas:
        for table in schema.tables:
            if table.data is None:
                continue
            for detail in table.data.sample_changed_detail:
                key_text = ", ".join(f"{k}={v}" for k, v in detail.primary_key.items())
                rows.append(
                    [
                        schema.schema_name,
                        table.table,
                        key_text,
                        detail.mismatch_column,
                        detail.source_value,
                        detail.target_value,
                        detail.source_row_hash,
                        detail.target_row_hash,
                        "Yes" if detail.verified else "No (row-number, unverified)",
                    ]
                )

    # Group by Source Schema, Source Table, Primary Key.
    rows.sort(key=lambda r: (r[0], r[1], r[2]))
    return rows


def _build_row_hash_rows(result: CatalogValidationResponse) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for schema in result.schemas:
        for table in schema.tables:
            if table.data is None:
                continue
            for mismatch in table.data.row_hash_mismatches:
                rows.append(
                    [
                        schema.schema_name,
                        table.table,
                        mismatch.primary_key,
                        mismatch.source_hash,
                        mismatch.target_hash,
                        mismatch.status,
                        mismatch.partition_bucket or "",
                    ]
                )

    # Group by Source Schema, Source Table, Primary Key.
    rows.sort(key=lambda r: (r[0], r[1], r[2]))

    # Prepend a 1-based sequential row number, assigned after sorting so it
    # reflects final display order (not discovery order).
    return [[i, *row] for i, row in enumerate(rows, start=1)]


def _build_suggestion_rows(result: CatalogValidationResponse) -> List[List[Any]]:
    """
    One plain-English sentence per issue found on a table, covering every
    category that can independently fail a table's Overall Status:
    schema/constraint issues (missing/extra column, column order, data
    type, nullable), per-column statistics (null count, distinct count,
    min/max), row count, and row-level data (pointing at the Data
    Mismatches / Row Hash Mismatches sheets for full detail rather than
    duplicating it here).

    If a table's row data actually matches (row count, row-hash/data
    comparison all passed) but the table still failed purely on
    schema/constraint or statistics grounds, an extra summary sentence
    calls that out explicitly, since it's an easy thing to miss buried in
    status columns. Every FAILed table gets at least one row here - if
    none of the specific categories below apply, a fallback sentence
    still points at Table Validation for the table's own Error field.
    """
    rows: List[List[Any]] = []

    for schema in result.schemas:
        for table in schema.tables:
            if table.status not in (ValidationStatus.FAIL, ValidationStatus.ERROR):
                continue

            table_issues: List[List[Any]] = []

            if table.schema_blocking:
                # BLOCKING schema difference (Tier 0) - the table was
                # aborted before any row-level tier ran, so there is no
                # row-hash/data finding to report alongside this. Surface
                # that explicitly rather than falling through to the
                # generic per-category checks below, most of which won't
                # have fired for a blocked table anyway.
                table_issues.append([
                    schema.schema_name, table.table, "-", "Blocked at Schema Check",
                    f"Table '{table.table}' has a schema difference severe enough "
                    f"(missing/extra column, incompatible data type, or a missing "
                    f"configured key column) that row-level comparison was not "
                    f"attempted - resolve the schema issue(s) below first, then "
                    f"re-run to check the data.",
                ])

            for col in table.missing_columns:
                table_issues.append([
                    schema.schema_name, table.table, col, "Missing Column",
                    f"Column '{col}' exists in the source but not in the target - "
                    f"it may not have been migrated, or was renamed/dropped.",
                ])

            for col in table.extra_columns:
                table_issues.append([
                    schema.schema_name, table.table, col, "Extra Column",
                    f"Column '{col}' exists in the target but not in the source - "
                    f"check whether it was added intentionally or is leftover from a prior load.",
                ])

            if table.column_order_status == ValidationStatus.FAIL:
                table_issues.append([
                    schema.schema_name, table.table, "-", "Column Order",
                    f"Columns match by name but are in a different order between source "
                    f"({', '.join(table.source_column_order)}) and target "
                    f"({', '.join(table.target_column_order)}).",
                ])

            if table.row_count_status == ValidationStatus.FAIL:
                diff = table.row_count_difference
                direction = "more" if (diff or 0) > 0 else "fewer"
                table_issues.append([
                    schema.schema_name, table.table, "-", "Row Count Mismatch",
                    f"Target has {abs(diff) if diff is not None else 'a different number of'} "
                    f"{direction} rows than source ({table.row_count_source} vs "
                    f"{table.row_count_target}) - check for a partial load, duplicate rows, "
                    f"or rows deleted/inserted after migration.",
                ])
            elif table.row_count_status == ValidationStatus.ERROR:
                table_issues.append([
                    schema.schema_name, table.table, "-", "Row Count Error",
                    f"Row count could not be verified for this table"
                    + (f": {table.error}" if table.error else " - see the Error column in Table Validation.")
                ])

            for col in table.columns:
                if col.data_type_status == ValidationStatus.FAIL:
                    table_issues.append([
                        schema.schema_name, table.table, col.column, "Data Type Mismatch",
                        f"Column '{col.column}' is {col.source_data_type} in the source but "
                        f"{col.target_data_type} in the target - if the row data still matches, "
                        f"only the declared type differs; if not, check for precision/format loss "
                        f"during migration.",
                    ])

                if col.nullable_status == ValidationStatus.FAIL:
                    src_null = "nullable" if col.source_nullable else "NOT NULL"
                    tgt_null = "nullable" if col.target_nullable else "NOT NULL"
                    table_issues.append([
                        schema.schema_name, table.table, col.column, "Nullable Mismatch",
                        f"Column '{col.column}' is declared {src_null} in the source but "
                        f"{tgt_null} in the target - this is a constraint difference, not "
                        f"necessarily a data problem, unless the stricter side is expected "
                        f"to reject values the other side allows.",
                    ])

                if col.null_count_status == ValidationStatus.FAIL:
                    table_issues.append([
                        schema.schema_name, table.table, col.column, "Null Count Mismatch",
                        f"Column '{col.column}' has {col.source_null_count} NULLs in the source "
                        f"but {col.target_null_count} in the target - some rows likely gained or "
                        f"lost a NULL value for this column during migration.",
                    ])

                if col.distinct_count_status == ValidationStatus.FAIL:
                    table_issues.append([
                        schema.schema_name, table.table, col.column, "Distinct Count Mismatch",
                        f"Column '{col.column}' has {col.source_distinct_count} distinct values "
                        f"in the source but {col.target_distinct_count} in the target - check for "
                        f"duplicate or collapsed values, or rows missing on one side.",
                    ])

                if col.min_max_status == ValidationStatus.FAIL:
                    table_issues.append([
                        schema.schema_name, table.table, col.column, "Min/Max Mismatch",
                        f"Column '{col.column}' ranges from {col.source_min} to {col.source_max} "
                        f"in the source but {col.target_min} to {col.target_max} in the target - "
                        f"check for outlier rows unique to one side, or a truncated/extended value range.",
                    ])

            data = table.data
            if data is not None:
                if data.row_hash_mismatch_count and data.row_hash_mismatch_count > 0:
                    by_row_number = data.key_columns == ["row_number"]
                    by_clause = (
                        "by row position (best-effort - no primary key "
                        "configured, so this cannot guarantee the same "
                        "record on both sides)"
                        if by_row_number
                        else "by primary key"
                    )
                    detail_clause = (
                        "values (marked unverified for row-position matches), if available"
                        if by_row_number
                        else "values, if available"
                    )
                    table_issues.append([
                        schema.schema_name, table.table, "-", "Row Data Mismatch",
                        f"{data.row_hash_mismatch_count} row(s) "
                        f"({data.row_hash_mismatch_percentage:.2f}%) differ between source and "
                        f"target {by_clause} - see the 'Row Hash Mismatches' sheet for the "
                        f"affected keys, and 'Data Mismatches' for the specific column(s) and "
                        f"{detail_clause}.",
                    ])
                elif data.status == ValidationStatus.FAIL and (
                    (data.source_only_rows or 0) > 0 or (data.target_only_rows or 0) > 0
                    or (data.changed_rows or 0) > 0
                ):
                    table_issues.append([
                        schema.schema_name, table.table, "-", "Row Data Mismatch",
                        f"Row-level comparison found {data.source_only_rows or 0} row(s) only in "
                        f"the source, {data.target_only_rows or 0} only in the target, and "
                        f"{data.changed_rows or 0} changed - see the 'Data Mismatches' sheet for detail.",
                    ])
                elif data.status == ValidationStatus.ERROR:
                    table_issues.append([
                        schema.schema_name, table.table, "-", "Row Data Comparison Error",
                        f"Row-level comparison could not complete for this table"
                        + (f": {data.error}" if data.error else " - see the Error column in Table Validation.")
                    ])

            if not table_issues:
                # Table FAILed/ERRORed but none of the categories above
                # explain why (e.g. an ERROR raised before any stage ran) -
                # never leave a failed table unexplained.
                table_issues.append([
                    schema.schema_name, table.table, "-", "Unclassified",
                    f"Table '{table.table}' has status {table.status.value} but the reason "
                    f"doesn't match a known category here - check the Error column and "
                    f"per-check status columns on the 'Table Validation' sheet for this table.",
                ])

            schema_only_issue = (
                not table.schema_blocking
                and table.status == ValidationStatus.FAIL
                and table.row_count_status != ValidationStatus.FAIL
                and (table.data is None or table.data.status != ValidationStatus.FAIL)
                and not any(issue[3] == "Row Data Mismatch" for issue in table_issues)
            )
            if schema_only_issue:
                rows.append([
                    schema.schema_name, table.table, "-", "Summary",
                    f"Table '{table.table}' failed, but the row data matches "
                    f"(row counts and row-hash/data comparison passed) - the failure "
                    f"is caused entirely by the schema/constraint issue(s) below.",
                ])
            rows.extend(table_issues)

    rows.sort(key=lambda r: (r[0], r[1]))
    return rows


def _build_summary_metrics(result: CatalogValidationResponse) -> List[Tuple[str, Any]]:
    total = passed = failed = errors = skipped = 0
    for schema in result.schemas:
        for table in schema.tables:
            total += 1
            if table.status == ValidationStatus.PASS:
                passed += 1
            elif table.status == ValidationStatus.ERROR:
                errors += 1
            elif table.status == ValidationStatus.SKIPPED:
                skipped += 1
            else:
                failed += 1
        total += len(schema.missing_tables)
        failed += len(schema.missing_tables)

    pass_pct = f"{(passed / total) * 100:.2f}%" if total else "0.00%"

    return [
        ("Total Tables", total),
        ("Passed Tables", passed),
        ("Failed Tables", failed),
        ("Error Tables", errors),
        ("Skipped Tables", skipped),
        ("Pass Percentage", pass_pct),
    ]


# ---------------------------------------------------------------------------
# CSV (Table Validation data only)
# ---------------------------------------------------------------------------
def generate_csv_report(
    result: CatalogValidationResponse,
    output_path: str,
) -> str:
    """
    Render the per-table validation results of a CatalogValidationResponse
    as a flat .csv file. Returns the output_path for convenience.
    """
    logger.info(
        "Generating CSV report | source=%s | target=%s | -> %s",
        result.source_catalog, result.target_catalog, output_path,
    )

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(TABLE_HEADERS)
        writer.writerows(_build_table_rows(result))

    logger.info("CSV report written to %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
def _write_header_row(ws: Worksheet, headers: List[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.freeze_panes = ws.cell(row=2, column=1)


def _write_rows(
    ws: Worksheet,
    rows: List[List[Any]],
    status_columns: set,
    group_col: Optional[int] = None,
    custom_status_map: Optional[dict] = None,
) -> int:
    """
    Write rows starting at row 2. If group_col is set, consecutive rows
    sharing that column's value are put into a collapsible outline group
    (Excel row grouping) - one level, matching the schema boundary.
    Returns the last row index written (1 if no data rows).
    """
    row_idx = 2
    prev_group_value = None
    group_start = None

    def _close_group(end_row: int) -> None:
        if group_start is not None and end_row > group_start:
            for r in range(group_start, end_row):
                ws.row_dimensions[r].outline_level = 1

    for values in rows:
        if group_col is not None:
            current_value = values[group_col - 1]
            if current_value != prev_group_value:
                if prev_group_value is not None:
                    _close_group(row_idx)
                prev_group_value = current_value
                group_start = row_idx + 1  # first detail row after the group's own header row

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if col_idx in status_columns:
                status = next((s for s in ValidationStatus if s.value == value), None)
                if status is None and custom_status_map is not None:
                    status = custom_status_map.get(value)
                if status in STATUS_FILLS:
                    cell.fill = STATUS_FILLS[status]
                    cell.font = STATUS_FONTS[status]
                elif value in ("MISSING_FROM_TARGET", "EXTRA_IN_TARGET"):
                    cell.fill = STATUS_FILLS[ValidationStatus.FAIL]
                    cell.font = STATUS_FONTS[ValidationStatus.FAIL]

        row_idx += 1

    if group_col is not None:
        _close_group(row_idx)

    return row_idx - 1


def _autofit(ws: Worksheet, headers: List[str], rows: List[List[Any]]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_values in rows:
            value = row_values[col_idx - 1]
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def _enable_filter(ws: Worksheet, num_cols: int, last_row: int) -> None:
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{max(last_row, 1)}"


# ---------------------------------------------------------------------------
# Excel sheet builders
# ---------------------------------------------------------------------------
def _build_summary_sheet(
    wb: Workbook,
    result: CatalogValidationResponse,
    source_type: Optional[str] = None,
    validations_run: Optional[str] = None,
) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Databricks Catalog Validation Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    ws["A2"] = f"Source: {result.source_catalog}   ->   Target: {result.target_catalog}"
    ws["A2"].font = VALUE_FONT
    ws.merge_cells("A2:B2")

    row = 4
    fields = [("Overall Status", _status_value(result.status))]
    if source_type:
        fields.append(("Source Type", source_type))
    if validations_run:
        fields.append(("Validations Run", validations_run))
    fields += [
        ("Validation Timestamp", result.validation_timestamp or ""),
        ("Duration (s)", result.execution_time_seconds),
    ]
    for label, value in fields:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = VALUE_FONT
        if label == "Overall Status":
            status = next((s for s in ValidationStatus if s.value == value), None)
            if status in STATUS_FILLS:
                cell.fill = STATUS_FILLS[status]
                cell.font = STATUS_FONTS[status]
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Table Summary").font = Font(
        name=FONT_NAME, size=12, bold=True, color="1F4E78"
    )
    row += 1

    header_row = row
    ws.cell(row=header_row, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=header_row, column=1).fill = HEADER_FILL
    ws.cell(row=header_row, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=header_row, column=2).fill = HEADER_FILL
    row += 1

    for label, value in _build_summary_metrics(result):
        ws.cell(row=row, column=1, value=label).font = VALUE_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = VALUE_FONT
        cell.alignment = Alignment(horizontal="right")
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40


def _build_table_validation_sheet(
    wb: Workbook,
    result: CatalogValidationResponse,
    enabled_validations: Optional[set] = None,
) -> None:
    ws = wb.create_sheet("Table Validation")
    rows = _build_table_rows(result)
    headers, rows, status_columns = _filter_table_columns(
        TABLE_HEADERS, rows, enabled_validations
    )

    _write_header_row(ws, headers)
    last_row = _write_rows(ws, rows, status_columns, group_col=1)
    _autofit(ws, headers, rows)
    _enable_filter(ws, len(headers), last_row)


def _build_column_validation_sheet(wb: Workbook, result: CatalogValidationResponse) -> None:
    ws = wb.create_sheet("Column Validation")
    rows = _build_column_rows(result)

    _write_header_row(ws, COLUMN_HEADERS)
    last_row = _write_rows(ws, rows, _COLUMN_STATUS_COLUMNS, group_col=1)
    _autofit(ws, COLUMN_HEADERS, rows)
    _enable_filter(ws, len(COLUMN_HEADERS), last_row)


def _build_data_mismatches_sheet(wb: Workbook, result: CatalogValidationResponse) -> None:
    ws = wb.create_sheet("Data Mismatches")
    rows = _build_mismatch_rows(result)

    _write_header_row(ws, MISMATCH_HEADERS)
    last_row = _write_rows(ws, rows, set(), group_col=1)
    _autofit(ws, MISMATCH_HEADERS, rows)
    _enable_filter(ws, len(MISMATCH_HEADERS), last_row)


def _build_row_hash_mismatches_sheet(wb: Workbook, result: CatalogValidationResponse) -> None:
    ws = wb.create_sheet("Row Hash Mismatches")
    rows = _build_row_hash_rows(result)

    _write_header_row(ws, ROW_HASH_HEADERS)
    last_row = _write_rows(
        ws, rows, {_ROW_HASH_STATUS_COLUMN}, group_col=2,
        custom_status_map=_ROW_HASH_STATUS_FILL_MAP,
    )
    _autofit(ws, ROW_HASH_HEADERS, rows)
    _enable_filter(ws, len(ROW_HASH_HEADERS), last_row)


def _build_suggestions_sheet(wb: Workbook, result: CatalogValidationResponse) -> None:
    ws = wb.create_sheet("Suggestions")
    rows = _build_suggestion_rows(result)

    _write_header_row(ws, SUGGESTION_HEADERS)

    row_idx = 2
    for values in rows:
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == len(SUGGESTION_HEADERS)))
        row_idx += 1

    ws.column_dimensions[get_column_letter(1)].width = 18
    ws.column_dimensions[get_column_letter(2)].width = 20
    ws.column_dimensions[get_column_letter(3)].width = 16
    ws.column_dimensions[get_column_letter(4)].width = 18
    ws.column_dimensions[get_column_letter(5)].width = 90

    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUGGESTION_HEADERS))}{max(row_idx - 1, 1)}"


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def generate_excel_report(
    result: CatalogValidationResponse,
    output_path: str,
    source_type: Optional[str] = None,
    enabled_validations: Optional[set] = None,
    skip_category_summary: bool = False,
) -> str:
    """
    Render a CatalogValidationResponse as a formatted, multi-sheet .xlsx
    workbook: Summary, Table Validation, Column Validation, Data
    Mismatches, Row Hash Mismatches, Mismatch Categories, Suggestions.
    Returns the output_path for convenience.

    source_type (e.g. "databricks"/"azure_blob"/"azure_sql") is optional
    and purely cosmetic - shown on the Summary sheet next to Overall
    Status so it's clear at a glance what kind of source was compared,
    since the report format itself is identical regardless of source.

    enabled_validations (e.g. {"catalog", "row"}) is optional; when given,
    Table Validation columns and whole sheets belonging to a validation
    type NOT in this set are omitted entirely - "column" gates the
    Column Validation sheet plus the column-level columns on Table
    Validation, "row" gates Data Mismatches/Row Hash Mismatches/Mismatch
    Categories (the last one is derived entirely from the same row-level
    mismatch detail, so it's gated identically) plus the row-level
    columns on Table Validation. None means "show everything" (no
    filtering), matching prior behavior for any caller that doesn't pass
    it.

    skip_category_summary (default False) omits ONLY the Mismatch
    Categories sheet, independent of enabled_validations - every other
    sheet (Data Mismatches included) renders exactly as it did before the
    mismatch-categorization feature existed. This is the mechanism behind
    the CLI's `--skip-category-summary` flag: a way to always recover the
    pre-categorization report byte-for-byte, in case the classifier's
    output is ever undesired or found to be wrong for a given run,
    without needing to also give up row-level validation entirely (which
    enabled_validations={"row" not included} would otherwise force).
    """
    logger.info(
        "Generating Excel report | source=%s | target=%s | source_type=%s | "
        "enabled_validations=%s | skip_category_summary=%s | -> %s",
        result.source_catalog, result.target_catalog, source_type,
        enabled_validations, skip_category_summary, output_path,
    )

    validations_run = (
        ", ".join(sorted(enabled_validations)) if enabled_validations is not None else None
    )

    wb = Workbook()

    _build_summary_sheet(wb, result, source_type, validations_run)
    _build_table_validation_sheet(wb, result, enabled_validations)

    if enabled_validations is None or "column" in enabled_validations:
        _build_column_validation_sheet(wb, result)

    if enabled_validations is None or "row" in enabled_validations:
        _build_data_mismatches_sheet(wb, result)
        _build_row_hash_mismatches_sheet(wb, result)
        if not skip_category_summary:
            _build_mismatch_category_sheet(wb, result)

    _build_suggestions_sheet(wb, result)

    wb.save(output_path)

    logger.info("Excel report written to %s", output_path)
    return output_path
