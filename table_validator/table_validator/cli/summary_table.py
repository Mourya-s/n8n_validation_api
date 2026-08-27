"""
Shared summary-table rendering, used by both `tablevalidator validate`
(right after generating a report) and `tablevalidator report` (reading an
existing one back). Both commands print the exact same aggregate figures
that live on the Excel report's Summary sheet - overall status, per-table
totals, pass percentage, and which validation types were run - so the
console output and the file never disagree.

SummaryData is the single shape both entry points build before handing
off to print_summary_table(), so the rendering logic itself is never
duplicated between the "just generated a live result" and "read an
existing .xlsx back" cases.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from rich.console import Console
from rich.table import Table

from table_validator.models import CatalogValidationResponse
from table_validator.reports.excel_report import _build_summary_metrics


@dataclass
class SummaryData:
    overall_status: str
    total_tables: int
    passed_tables: int
    failed_tables: int
    error_tables: int
    skipped_tables: int
    pass_percentage: str
    source_type: Optional[str] = None
    validations_run: Optional[str] = None
    validation_timestamp: Optional[str] = None
    duration_seconds: Optional[float] = None


def summary_from_response(
    result: CatalogValidationResponse,
    source_type: Optional[str] = None,
    validations_run: Optional[str] = None,
) -> SummaryData:
    """Build a SummaryData from a freshly-computed CatalogValidationResponse,
    using the exact same _build_summary_metrics() the Excel Summary sheet
    itself is built from - so the console table and the file always agree."""
    metrics = dict(_build_summary_metrics(result))
    return SummaryData(
        overall_status=result.status.value,
        total_tables=metrics["Total Tables"],
        passed_tables=metrics["Passed Tables"],
        failed_tables=metrics["Failed Tables"],
        error_tables=metrics["Error Tables"],
        skipped_tables=metrics["Skipped Tables"],
        pass_percentage=metrics["Pass Percentage"],
        source_type=source_type,
        validations_run=validations_run,
        validation_timestamp=result.validation_timestamp,
        duration_seconds=result.execution_time_seconds,
    )


def summary_from_excel(path: Path) -> SummaryData:
    """
    Read a SummaryData back out of a saved report's Summary sheet.

    The Summary sheet's row layout shifts depending on which optional
    fields (Source Type, Validations Run) were present when it was
    generated, so this scans column A for each known label by name
    rather than assuming fixed row numbers.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb["Summary"]

        fields: dict = {}
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            label, value = row[0], row[1]
            if label is not None:
                fields[str(label)] = value
    finally:
        # read_only workbooks hold an open file handle until closed -
        # without this, a caller that immediately tries to overwrite or
        # delete the same path (e.g. a subsequent `validate` run, or a
        # test's tempdir cleanup) can hit a Windows file-lock error.
        wb.close()

    def _metric(label: str, default=0):
        value = fields.get(label, default)
        return int(value) if isinstance(value, (int, float)) else default

    return SummaryData(
        overall_status=str(fields.get("Overall Status", "UNKNOWN")),
        total_tables=_metric("Total Tables"),
        passed_tables=_metric("Passed Tables"),
        failed_tables=_metric("Failed Tables"),
        error_tables=_metric("Error Tables"),
        skipped_tables=_metric("Skipped Tables"),
        pass_percentage=str(fields.get("Pass Percentage", "0.00%")),
        source_type=fields.get("Source Type"),
        validations_run=fields.get("Validations Run"),
        validation_timestamp=fields.get("Validation Timestamp") or None,
        duration_seconds=fields.get("Duration (s)"),
    )


_STATUS_STYLES = {
    "PASS": "bold green",
    "FAIL": "bold red",
    "ERROR": "bold yellow",
    "SKIPPED": "bold white",
}


def print_summary_table(data: SummaryData, console: Optional[Console] = None) -> None:
    """Render a SummaryData as a compact rich Table on the console."""
    console = console or Console()

    status_style = _STATUS_STYLES.get(data.overall_status, "bold white")

    table = Table(title="Validation Summary", show_lines=False)
    table.add_column("Metric", style="bold")
    table.add_column("Value")

    table.add_row("Overall Status", f"[{status_style}]{data.overall_status}[/{status_style}]")
    if data.source_type:
        table.add_row("Source Type", data.source_type)
    if data.validations_run:
        table.add_row("Validations Run", data.validations_run)
    table.add_row("Total Tables", str(data.total_tables))
    table.add_row("Passed Tables", f"[green]{data.passed_tables}[/green]")
    table.add_row("Failed Tables", f"[red]{data.failed_tables}[/red]")
    table.add_row("Error Tables", f"[yellow]{data.error_tables}[/yellow]")
    table.add_row("Skipped Tables", str(data.skipped_tables))
    table.add_row("Pass Percentage", data.pass_percentage)
    if data.validation_timestamp:
        table.add_row("Validation Timestamp", data.validation_timestamp)
    if data.duration_seconds is not None:
        table.add_row("Duration (s)", f"{data.duration_seconds:.3f}")

    console.print(table)
