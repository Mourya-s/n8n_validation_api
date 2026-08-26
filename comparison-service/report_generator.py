"""
Report Generator (CSV + Excel)

Consumes a CatalogValidationResponse (produced by
comparison_engine.CatalogValidator) and renders it as either a flat .csv
file (Table Validation data only) or a formatted, multi-sheet .xlsx
workbook (Summary / Table Validation / Column Validation / Data
Mismatches / Row Hash Mismatches).

Deliberately kept out of comparison_engine.py per the project spec:
the validator returns clean structured data, and this module is the
only place that knows about report formatting (csv / openpyxl).
"""

from __future__ import annotations

import csv
import logging
from typing import Any, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models import CatalogValidationResponse, ValidationStatus

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Shared column definitions
# ---------------------------------------------------------------------------
TABLE_HEADERS = [
    "Source Schema",
    "Source Table",
    "Target Schema",
    "Target Table",
    "Overall Status",
    "Schema Match",
    "Column Order",
    "Row Count (Src)",
    "Row Count (Tgt)",
    "Row Count Diff",
    "Data Types",
    "Nullable",
    "Null Counts",
    "Distinct Counts",
    "Min/Max",
    "Data Match",
    "Mismatch Count",
    "Mismatch %",
    "Row Hash Mismatch Count",
    "Row Hash Mismatch %",
    "Validation Timestamp",
    "Duration",
]

# 1-based column indices within TABLE_HEADERS holding a PASS/FAIL/ERROR/SKIPPED value.
_TABLE_STATUS_COLUMNS = {5, 6, 7, 11, 12, 13, 14, 15, 16}

COLUMN_HEADERS = [
    "Source Schema", "Source Table", "Column", "Status",
    "Source Type", "Target Type", "Type Status",
    "Source Nullable", "Target Nullable", "Nullable Status",
    "Source Nulls", "Target Nulls", "Null Status",
    "Source Distinct", "Target Distinct", "Distinct Status",
    "Source Min", "Source Max", "Target Min", "Target Max", "Min/Max Status",
    "Error",
]
_COLUMN_STATUS_COLUMNS = {4, 7, 10, 13, 16, 21}

MISMATCH_HEADERS = [
    "Source Schema", "Source Table", "Primary Key",
    "Mismatch Column", "Source Value", "Target Value",
    "Row Hash (Source)", "Row Hash (Target)",
]

ROW_HASH_HEADERS = [
    "Row #", "Source Schema", "Source Table",
    "Primary Key", "Source Hash", "Target Hash", "Mismatch Status",
]
_ROW_HASH_STATUS_COLUMN = 7

SUGGESTION_HEADERS = [
    "Source Schema", "Source Table", "Column", "Issue Type", "Suggestion",
]

SUMMARY_METRIC_LABELS = [
    "Total Tables", "Passed Tables", "Failed Tables",
    "Error Tables", "Skipped Tables", "Pass Percentage",
]

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="1F4E78")
LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True)
VALUE_FONT = Font(name=FONT_NAME, size=10)
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

STATUS_FILLS = {
    ValidationStatus.PASS: PatternFill("solid", fgColor="C6EFCE"),
    ValidationStatus.FAIL: PatternFill("solid", fgColor="FFC7CE"),
    ValidationStatus.ERROR: PatternFill("solid", fgColor="FFD8A8"),
    ValidationStatus.SKIPPED: PatternFill("solid", fgColor="E7E6E6"),
}
STATUS_FONTS = {
    ValidationStatus.PASS: Font(name=FONT_NAME, size=10, bold=True, color="006100"),
    ValidationStatus.FAIL: Font(name=FONT_NAME, size=10, bold=True, color="9C0006"),
    ValidationStatus.ERROR: Font(name=FONT_NAME, size=10, bold=True, color="974706"),
    ValidationStatus.SKIPPED: Font(name=FONT_NAME, size=10, bold=True, color="666666"),
}

# Sort rank so FAIL/ERROR surface above PASS/SKIPPED within a schema.
_STATUS_SORT_RANK = {
    "FAIL": 0,
    "ERROR": 0,
    "MISSING_FROM_TARGET": 0,
    "PASS": 1,
    "SKIPPED": 1,
}

# Row-hash mismatch statuses aren't ValidationStatus members - map them onto
# the same PASS/FAIL/ERROR/SKIPPED fill+font conventions for the Row Hash
# Mismatches sheet (all three outcomes here are a real difference -> FAIL).
_ROW_HASH_STATUS_FILL_MAP = {
    "MISMATCH": ValidationStatus.FAIL,
    "MISSING_IN_TARGET": ValidationStatus.FAIL,
    "MISSING_IN_SOURCE": ValidationStatus.FAIL,
}


def _status_value(status: Optional[ValidationStatus]) -> str:
    if status is None:
        return ""
    return status.value if isinstance(status, ValidationStatus) else str(status)


def _mismatch_count(table) -> Optional[int]:
    if table.data is None:
        return None
    counts = [table.data.source_only_rows, table.data.target_only_rows, table.data.changed_rows]
    if all(c is None for c in counts):
        return None
    return sum(c or 0 for c in counts)


def _mismatch_pct(table, mismatch_count: Optional[int]) -> str:
    if mismatch_count is None:
        return ""
    total = table.row_count_source or 0
    if total <= 0:
        return ""
    return f"{(mismatch_count / total) * 100:.2f}%"


# ---------------------------------------------------------------------------
# Row builders (shared by CSV + every Excel sheet)
# ---------------------------------------------------------------------------
def _build_table_rows(result: CatalogValidationResponse) -> List[List[Any]]:
    timestamp = result.validation_timestamp or ""
    duration = result.execution_time_seconds

    rows: List[List[Any]] = []
    for schema in result.schemas:
        for table in schema.tables:
            data_status = table.data.status if table.data else None
            mismatch_count = _mismatch_count(table)

            row_hash_count = table.data.row_hash_mismatch_count if table.data else 0
            row_hash_pct = (
                f"{table.data.row_hash_mismatch_percentage:.2f}%" if table.data else ""
            )

            rows.append(
                [
                    schema.schema_name,
                    table.table,
                    schema.schema_name,
                    table.table,
                    _status_value(table.status),
                    _status_value(table.columns_status),
                    _status_value(table.column_order_status),
                    table.row_count_source,
                    table.row_count_target,
                    table.row_count_difference,
                    _status_value(table.data_types_status),
                    _status_value(table.nullable_status),
                    _status_value(table.null_counts_status),
                    _status_value(table.distinct_counts_status),
                    _status_value(table.min_max_status),
                    _status_value(data_status),
                    mismatch_count if mismatch_count is not None else "",
                    _mismatch_pct(table, mismatch_count),
                    row_hash_count,
                    row_hash_pct,
                    timestamp,
                    duration,
                ]
            )

        for missing_table in schema.missing_tables:
            rows.append(
                [
                    schema.schema_name, missing_table, schema.schema_name, missing_table,
                    "MISSING_FROM_TARGET",
                ]
                + [""] * 15
                + [timestamp, duration]
            )

    # Sort: Source Schema asc, then FAIL/ERROR before PASS/SKIPPED, then Source Table asc.
    rows.sort(key=lambda r: (r[0], _STATUS_SORT_RANK.get(r[4], 1), r[1]))
    return rows


def _build_column_rows(result: CatalogValidationResponse) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for schema in result.schemas:
        for table in schema.tables:
            for col in table.columns:
                rows.append(
                    [
                        schema.schema_name, table.table, col.column, _status_value(col.status),
                        col.source_data_type, col.target_data_type, _status_value(col.data_type_status),
                        col.source_nullable, col.target_nullable, _status_value(col.nullable_status),
                        col.source_null_count, col.target_null_count, _status_value(col.null_count_status),
                        col.source_distinct_count, col.target_distinct_count, _status_value(col.distinct_count_status),
                        col.source_min, col.source_max, col.target_min, col.target_max,
                        _status_value(col.min_max_status),
                        col.error or "",
                    ]
                )

            for missing_col in table.missing_columns:
                rows.append(
                    [schema.schema_name, table.table, missing_col, "MISSING_FROM_TARGET"]
                    + [""] * 18
                )
            for extra_col in table.extra_columns:
                rows.append(
                    [schema.schema_name, table.table, extra_col, "EXTRA_IN_TARGET"]
                    + [""] * 18
                )

    rows.sort(key=lambda r: (r[0], r[1], r[2]))
    return rows


def _build_mismatch_rows(result: CatalogValidationResponse) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for schema in result.schemas:
        for table in schema.tables:
            if table.data is None:
                continue
            for detail in table.data.sample_changed_detail:
                key_text = ", ".join(f"{k}={v}" for k, v in detail.primary_key.items())
                rows.append(
                    [
                        schema.schema_name,
                        table.table,
                        key_text,
                        detail.mismatch_column,
                        detail.source_value,
                        detail.target_value,
                        detail.source_row_hash,
                        detail.target_row_hash,
                    ]
                )

    # Group by Source Schema, Source Table, Primary Key.
    rows.sort(key=lambda r: (r[0], r[1], r[2]))
    return rows


def _build_row_hash_rows(result: CatalogValidationResponse) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for schema in result.schemas:
        for table in schema.tables:
            if table.data is None:
                continue
            for mismatch in table.data.row_hash_mismatches:
                rows.append(
                    [
                        schema.schema_name,
                        table.table,
                        mismatch.primary_key,
                        mismatch.source_hash,
                        mismatch.target_hash,
                        mismatch.status,
                    ]
                )

    # Group by Source Schema, Source Table, Primary Key.
    rows.sort(key=lambda r: (r[0], r[1], r[2]))

    # Prepend a 1-based sequential row number, assigned after sorting so it
    # reflects final display order (not discovery order).
    return [[i, *row] for i, row in enumerate(rows, start=1)]


def _build_suggestion_rows(result: CatalogValidationResponse) -> List[List[Any]]:
    """
    One plain-English sentence per issue found on a table, covering every
    category that can independently fail a table's Overall Status:
    schema/constraint issues (missing/extra column, column order, data
    type, nullable), per-column statistics (null count, distinct count,
    min/max), row count, and row-level data (pointing at the Data
    Mismatches / Row Hash Mismatches sheets for full detail rather than
    duplicating it here).

    If a table's row data actually matches (row count, row-hash/data
    comparison all passed) but the table still failed purely on
    schema/constraint or statistics grounds, an extra summary sentence
    calls that out explicitly, since it's an easy thing to miss buried in
    status columns. Every FAILed table gets at least one row here - if
    none of the specific categories below apply, a fallback sentence
    still points at Table Validation for the table's own Error field.
    """
    rows: List[List[Any]] = []

    for schema in result.schemas:
        for table in schema.tables:
            if table.status not in (ValidationStatus.FAIL, ValidationStatus.ERROR):
                continue

            table_issues: List[List[Any]] = []

            for col in table.missing_columns:
                table_issues.append([
                    schema.schema_name, table.table, col, "Missing Column",
                    f"Column '{col}' exists in the source but not in the target - "
                    f"it may not have been migrated, or was renamed/dropped.",
                ])

            for col in table.extra_columns:
                table_issues.append([
                    schema.schema_name, table.table, col, "Extra Column",
                    f"Column '{col}' exists in the target but not in the source - "
                    f"check whether it was added intentionally or is leftover from a prior load.",
                ])

            if table.column_order_status == ValidationStatus.FAIL:
                table_issues.append([
                    schema.schema_name, table.table, "-", "Column Order",
                    f"Columns match by name but are in a different order between source "
                    f"({', '.join(table.source_column_order)}) and target "
                    f"({', '.join(table.target_column_order)}).",
                ])

            if table.row_count_status == ValidationStatus.FAIL:
                diff = table.row_count_difference
                direction = "more" if (diff or 0) > 0 else "fewer"
                table_issues.append([
                    schema.schema_name, table.table, "-", "Row Count Mismatch",
                    f"Target has {abs(diff) if diff is not None else 'a different number of'} "
                    f"{direction} rows than source ({table.row_count_source} vs "
                    f"{table.row_count_target}) - check for a partial load, duplicate rows, "
                    f"or rows deleted/inserted after migration.",
                ])
            elif table.row_count_status == ValidationStatus.ERROR:
                table_issues.append([
                    schema.schema_name, table.table, "-", "Row Count Error",
                    f"Row count could not be verified for this table"
                    + (f": {table.error}" if table.error else " - see the Error column in Table Validation.")
                ])

            for col in table.columns:
                if col.data_type_status == ValidationStatus.FAIL:
                    table_issues.append([
                        schema.schema_name, table.table, col.column, "Data Type Mismatch",
                        f"Column '{col.column}' is {col.source_data_type} in the source but "
                        f"{col.target_data_type} in the target - if the row data still matches, "
                        f"only the declared type differs; if not, check for precision/format loss "
                        f"during migration.",
                    ])

                if col.nullable_status == ValidationStatus.FAIL:
                    src_null = "nullable" if col.source_nullable else "NOT NULL"
                    tgt_null = "nullable" if col.target_nullable else "NOT NULL"
                    table_issues.append([
                        schema.schema_name, table.table, col.column, "Nullable Mismatch",
                        f"Column '{col.column}' is declared {src_null} in the source but "
                        f"{tgt_null} in the target - this is a constraint difference, not "
                        f"necessarily a data problem, unless the stricter side is expected "
                        f"to reject values the other side allows.",
                    ])

                if col.null_count_status == ValidationStatus.FAIL:
                    table_issues.append([
                        schema.schema_name, table.table, col.column, "Null Count Mismatch",
                        f"Column '{col.column}' has {col.source_null_count} NULLs in the source "
                        f"but {col.target_null_count} in the target - some rows likely gained or "
                        f"lost a NULL value for this column during migration.",
                    ])

                if col.distinct_count_status == ValidationStatus.FAIL:
                    table_issues.append([
                        schema.schema_name, table.table, col.column, "Distinct Count Mismatch",
                        f"Column '{col.column}' has {col.source_distinct_count} distinct values "
                        f"in the source but {col.target_distinct_count} in the target - check for "
                        f"duplicate or collapsed values, or rows missing on one side.",
                    ])

                if col.min_max_status == ValidationStatus.FAIL:
                    table_issues.append([
                        schema.schema_name, table.table, col.column, "Min/Max Mismatch",
                        f"Column '{col.column}' ranges from {col.source_min} to {col.source_max} "
                        f"in the source but {col.target_min} to {col.target_max} in the target - "
                        f"check for outlier rows unique to one side, or a truncated/extended value range.",
                    ])

            data = table.data
            if data is not None:
                if data.row_hash_mismatch_count and data.row_hash_mismatch_count > 0:
                    table_issues.append([
                        schema.schema_name, table.table, "-", "Row Data Mismatch",
                        f"{data.row_hash_mismatch_count} row(s) "
                        f"({data.row_hash_mismatch_percentage:.2f}%) differ between source and "
                        f"target by primary key - see the 'Row Hash Mismatches' sheet for the "
                        f"affected keys, and 'Data Mismatches' for the specific column(s) and "
                        f"values, if available.",
                    ])
                elif data.status == ValidationStatus.FAIL and (
                    (data.source_only_rows or 0) > 0 or (data.target_only_rows or 0) > 0
                    or (data.changed_rows or 0) > 0
                ):
                    table_issues.append([
                        schema.schema_name, table.table, "-", "Row Data Mismatch",
                        f"Row-level comparison found {data.source_only_rows or 0} row(s) only in "
                        f"the source, {data.target_only_rows or 0} only in the target, and "
                        f"{data.changed_rows or 0} changed - see the 'Data Mismatches' sheet for detail.",
                    ])
                elif data.status == ValidationStatus.ERROR:
                    table_issues.append([
                        schema.schema_name, table.table, "-", "Row Data Comparison Error",
                        f"Row-level comparison could not complete for this table"
                        + (f": {data.error}" if data.error else " - see the Error column in Table Validation.")
                    ])

            if not table_issues:
                # Table FAILed/ERRORed but none of the categories above
                # explain why (e.g. an ERROR raised before any stage ran) -
                # never leave a failed table unexplained.
                table_issues.append([
                    schema.schema_name, table.table, "-", "Unclassified",
                    f"Table '{table.table}' has status {table.status.value} but the reason "
                    f"doesn't match a known category here - check the Error column and "
                    f"per-check status columns on the 'Table Validation' sheet for this table.",
                ])

            schema_only_issue = (
                table.status == ValidationStatus.FAIL
                and table.row_count_status != ValidationStatus.FAIL
                and (table.data is None or table.data.status != ValidationStatus.FAIL)
                and not any(issue[3] == "Row Data Mismatch" for issue in table_issues)
            )
            if schema_only_issue:
                rows.append([
                    schema.schema_name, table.table, "-", "Summary",
                    f"Table '{table.table}' failed, but the row data matches "
                    f"(row counts and row-hash/data comparison passed) - the failure "
                    f"is caused entirely by the schema/constraint issue(s) below.",
                ])
            rows.extend(table_issues)

    rows.sort(key=lambda r: (r[0], r[1]))
    return rows


def _build_summary_metrics(result: CatalogValidationResponse) -> List[Tuple[str, Any]]:
    total = passed = failed = errors = skipped = 0
    for schema in result.schemas:
        for table in schema.tables:
            total += 1
            if table.status == ValidationStatus.PASS:
                passed += 1
            elif table.status == ValidationStatus.ERROR:
                errors += 1
            elif table.status == ValidationStatus.SKIPPED:
                skipped += 1
            else:
                failed += 1
        total += len(schema.missing_tables)
        failed += len(schema.missing_tables)

    pass_pct = f"{(passed / total) * 100:.2f}%" if total else "0.00%"

    return [
        ("Total Tables", total),
        ("Passed Tables", passed),
        ("Failed Tables", failed),
        ("Error Tables", errors),
        ("Skipped Tables", skipped),
        ("Pass Percentage", pass_pct),
    ]


# ---------------------------------------------------------------------------
# CSV (Table Validation data only)
# ---------------------------------------------------------------------------
def generate_csv_report(
    result: CatalogValidationResponse,
    output_path: str,
) -> str:
    """
    Render the per-table validation results of a CatalogValidationResponse
    as a flat .csv file. Returns the output_path for convenience.
    """
    logger.info(
        "Generating CSV report | source=%s | target=%s | -> %s",
        result.source_catalog, result.target_catalog, output_path,
    )

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(TABLE_HEADERS)
        writer.writerows(_build_table_rows(result))

    logger.info("CSV report written to %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
def _write_header_row(ws: Worksheet, headers: List[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.freeze_panes = ws.cell(row=2, column=1)


def _write_rows(
    ws: Worksheet,
    rows: List[List[Any]],
    status_columns: set,
    group_col: Optional[int] = None,
    custom_status_map: Optional[dict] = None,
) -> int:
    """
    Write rows starting at row 2. If group_col is set, consecutive rows
    sharing that column's value are put into a collapsible outline group
    (Excel row grouping) - one level, matching the schema boundary.
    Returns the last row index written (1 if no data rows).
    """
    row_idx = 2
    prev_group_value = None
    group_start = None

    def _close_group(end_row: int) -> None:
        if group_start is not None and end_row > group_start:
            for r in range(group_start, end_row):
                ws.row_dimensions[r].outline_level = 1

    for values in rows:
        if group_col is not None:
            current_value = values[group_col - 1]
            if current_value != prev_group_value:
                if prev_group_value is not None:
                    _close_group(row_idx)
                prev_group_value = current_value
                group_start = row_idx + 1  # first detail row after the group's own header row

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if col_idx in status_columns:
                status = next((s for s in ValidationStatus if s.value == value), None)
                if status is None and custom_status_map is not None:
                    status = custom_status_map.get(value)
                if status in STATUS_FILLS:
                    cell.fill = STATUS_FILLS[status]
                    cell.font = STATUS_FONTS[status]
                elif value in ("MISSING_FROM_TARGET", "EXTRA_IN_TARGET"):
                    cell.fill = STATUS_FILLS[ValidationStatus.FAIL]
                    cell.font = STATUS_FONTS[ValidationStatus.FAIL]

        row_idx += 1

    if group_col is not None:
        _close_group(row_idx)

    return row_idx - 1


def _autofit(ws: Worksheet, headers: List[str], rows: List[List[Any]]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_values in rows:
            value = row_values[col_idx - 1]
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def _enable_filter(ws: Worksheet, num_cols: int, last_row: int) -> None:
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{max(last_row, 1)}"


# ---------------------------------------------------------------------------
# Excel sheet builders
# ---------------------------------------------------------------------------
def _build_summary_sheet(wb: Workbook, result: CatalogValidationResponse) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Databricks Catalog Validation Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    ws["A2"] = f"Source: {result.source_catalog}   ->   Target: {result.target_catalog}"
    ws["A2"].font = VALUE_FONT
    ws.merge_cells("A2:B2")

    row = 4
    for label, value in [
        ("Overall Status", _status_value(result.status)),
        ("Validation Timestamp", result.validation_timestamp or ""),
        ("Duration (s)", result.execution_time_seconds),
    ]:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = VALUE_FONT
        if label == "Overall Status":
            status = next((s for s in ValidationStatus if s.value == value), None)
            if status in STATUS_FILLS:
                cell.fill = STATUS_FILLS[status]
                cell.font = STATUS_FONTS[status]
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Table Summary").font = Font(
        name=FONT_NAME, size=12, bold=True, color="1F4E78"
    )
    row += 1

    header_row = row
    ws.cell(row=header_row, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=header_row, column=1).fill = HEADER_FILL
    ws.cell(row=header_row, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=header_row, column=2).fill = HEADER_FILL
    row += 1

    for label, value in _build_summary_metrics(result):
        ws.cell(row=row, column=1, value=label).font = VALUE_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = VALUE_FONT
        cell.alignment = Alignment(horizontal="right")
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40


def _build_table_validation_sheet(wb: Workbook, result: CatalogValidationResponse) -> None:
    ws = wb.create_sheet("Table Validation")
    rows = _build_table_rows(result)

    _write_header_row(ws, TABLE_HEADERS)
    last_row = _write_rows(ws, rows, _TABLE_STATUS_COLUMNS, group_col=1)
    _autofit(ws, TABLE_HEADERS, rows)
    _enable_filter(ws, len(TABLE_HEADERS), last_row)


def _build_column_validation_sheet(wb: Workbook, result: CatalogValidationResponse) -> None:
    ws = wb.create_sheet("Column Validation")
    rows = _build_column_rows(result)

    _write_header_row(ws, COLUMN_HEADERS)
    last_row = _write_rows(ws, rows, _COLUMN_STATUS_COLUMNS, group_col=1)
    _autofit(ws, COLUMN_HEADERS, rows)
    _enable_filter(ws, len(COLUMN_HEADERS), last_row)


def _build_data_mismatches_sheet(wb: Workbook, result: CatalogValidationResponse) -> None:
    ws = wb.create_sheet("Data Mismatches")
    rows = _build_mismatch_rows(result)

    _write_header_row(ws, MISMATCH_HEADERS)
    last_row = _write_rows(ws, rows, set(), group_col=1)
    _autofit(ws, MISMATCH_HEADERS, rows)
    _enable_filter(ws, len(MISMATCH_HEADERS), last_row)


def _build_row_hash_mismatches_sheet(wb: Workbook, result: CatalogValidationResponse) -> None:
    ws = wb.create_sheet("Row Hash Mismatches")
    rows = _build_row_hash_rows(result)

    _write_header_row(ws, ROW_HASH_HEADERS)
    last_row = _write_rows(
        ws, rows, {_ROW_HASH_STATUS_COLUMN}, group_col=2,
        custom_status_map=_ROW_HASH_STATUS_FILL_MAP,
    )
    _autofit(ws, ROW_HASH_HEADERS, rows)
    _enable_filter(ws, len(ROW_HASH_HEADERS), last_row)


def _build_suggestions_sheet(wb: Workbook, result: CatalogValidationResponse) -> None:
    ws = wb.create_sheet("Suggestions")
    rows = _build_suggestion_rows(result)

    _write_header_row(ws, SUGGESTION_HEADERS)

    row_idx = 2
    for values in rows:
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == len(SUGGESTION_HEADERS)))
        row_idx += 1

    ws.column_dimensions[get_column_letter(1)].width = 18
    ws.column_dimensions[get_column_letter(2)].width = 20
    ws.column_dimensions[get_column_letter(3)].width = 16
    ws.column_dimensions[get_column_letter(4)].width = 18
    ws.column_dimensions[get_column_letter(5)].width = 90

    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUGGESTION_HEADERS))}{max(row_idx - 1, 1)}"


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def generate_excel_report(
    result: CatalogValidationResponse,
    output_path: str,
) -> str:
    """
    Render a CatalogValidationResponse as a formatted, multi-sheet .xlsx
    workbook: Summary, Table Validation, Column Validation, Data
    Mismatches, Row Hash Mismatches, Suggestions. Returns the output_path
    for convenience.
    """
    logger.info(
        "Generating Excel report | source=%s | target=%s | -> %s",
        result.source_catalog, result.target_catalog, output_path,
    )

    wb = Workbook()

    _build_summary_sheet(wb, result)
    _build_table_validation_sheet(wb, result)
    _build_column_validation_sheet(wb, result)
    _build_data_mismatches_sheet(wb, result)
    _build_row_hash_mismatches_sheet(wb, result)
    _build_suggestions_sheet(wb, result)

    wb.save(output_path)

    logger.info("Excel report written to %s", output_path)
    return output_path
